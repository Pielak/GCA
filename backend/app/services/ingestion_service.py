"""
Ingestion Service — Upload, deduplicação e gestão de documentos por projeto.
"""
import asyncio
import hashlib
import io
import json
import re
from datetime import datetime, timezone
from typing import Optional
from uuid import uuid4, UUID

from sqlalchemy import select, func, and_
from sqlalchemy.ext.asyncio import AsyncSession
import structlog

from app.models.base import IngestedDocument, ArguiderAnalysis, ModuleCandidate, OCG, OCGDeltaLog
from app.services.arguider_service import ArguiderService, DocumentExtractor

logger = structlog.get_logger(__name__)

# Tipos de arquivo aceitos → file_type
# B2 (auditoria 2026-05-04): adicionados html/htm, rtf, eml — extratores
# em _dispatch_to_n8n via libs html2text/striprtf/email (built-in).
EXTENSION_MAP = {
    "pdf": "pdf", "docx": "docx", "doc": "docx",
    "md": "markdown", "txt": "markdown",
    "html": "html", "htm": "html",
    "rtf": "rtf",
    "eml": "eml",
    "png": "image", "jpg": "image", "jpeg": "image", "gif": "image", "webp": "image",
    "xlsx": "spreadsheet", "xls": "spreadsheet", "csv": "spreadsheet",
    "py": "code", "ts": "code", "js": "code", "java": "code",
    "cs": "code", "go": "code", "rs": "code",
    # MVP 25 Fase 25.3 — stylesheets alimentam o extractor determinístico
    # de design tokens (Fase 25.1). Não vão pro pipeline LLM.
    "css": "stylesheet", "scss": "stylesheet",
}

MAX_FILE_SIZE = 50 * 1024 * 1024  # 50MB


_M01_MARKER_RE = re.compile(r"gca_iteration_id=([0-9a-fA-F-]{36})")


def _extract_m01_iteration_marker(file_bytes: bytes, file_type: str) -> "UUID | None":
    """M01 — extrai o UUID da iteração do metadado `Keywords` do documento.

    O gerador M01 carimba `keywords = gca_iteration_id=<uuid>` nos metadados
    canônicos — PDF via `canvas.setKeywords(...)`, DOCX via
    `core_properties.keywords`. Ambos sobrevivem a save/edit nos editores
    comuns. Retorna None se o marker não existe ou é inválido.
    """
    try:
        keywords_raw = ""
        if file_type == "pdf":
            import pdfplumber
            with pdfplumber.open(io.BytesIO(file_bytes)) as pdf:
                meta = pdf.metadata or {}
                keywords_raw = str(meta.get("Keywords") or "")
        elif file_type == "docx":
            from docx import Document
            doc = Document(io.BytesIO(file_bytes))
            keywords_raw = str(doc.core_properties.keywords or "")
        else:
            return None

        if not keywords_raw:
            return None
        m = _M01_MARKER_RE.search(keywords_raw)
        if not m:
            return None
        from uuid import UUID as _UUID
        return _UUID(m.group(1))
    except Exception:  # noqa: BLE001
        return None


async def _propagate_async(
    project_id: UUID,
    changes: list[dict],
    ocg_version: Optional[int],
) -> None:
    """Roda PropagationService em sessão própria, fire-and-forget.

    Isolamento: se a propagação falhar (ex: backlog regen quebra), o erro
    NÃO afeta a transação do OCG (já commitada). Apenas loga.

    Sessão dedicada: não compartilha com a sessão da ingestão (que pode
    estar fechando). Cada propagation tem seu próprio ciclo de vida.
    """
    try:
        from app.db.database import AsyncSessionLocal
        from app.services.propagation_service import PropagationService

        async with AsyncSessionLocal() as db:
            propagator = PropagationService(db)
            await propagator.propagate(
                project_id=project_id,
                changes=changes,
                ocg_version=ocg_version,
            )
    except Exception as exc:  # noqa: BLE001
        import traceback
        logger.warning(
            "ingestion.propagate_async_failed",
            project_id=str(project_id),
            error=str(exc) or repr(exc),
            error_type=type(exc).__name__,
            traceback=traceback.format_exc(),
        )


async def dispatch_first_pending_for_project(
    db: AsyncSession,
    project_id: UUID,
) -> Optional[UUID]:
    """Dispatcher canônico de fila por projeto (1-por-vez, n8n ou Celery).

    Sob advisory lock Postgres por project_id (atômico contra uploads
    paralelos):
      1. Se já existe doc 'processing' (não-deleted), nada a fazer.
      2. Senão, pega o doc 'pending' mais antigo, marca 'processing' e
         dispara o pipeline (n8n se INGESTION_VIA_N8N, senão Celery).

    Chamado em 2 lugares:
      - upload_document → puxa o que acabou de chegar.
      - webhook /ingestion-complete → puxa o próximo da fila quando o
        anterior termina.

    Retorna o UUID do doc despachado, ou None se nada a fazer.
    """
    from sqlalchemy import text as _text
    from app.core.config import settings
    from app.models.base import IngestedDocument as _IngDoc
    from app.tasks.pipeline import pipeline_ingest_task

    # Advisory lock por project_id — serializa concorrência sem locking de tabela.
    # Métricas de contenção (DBA F1 R3): warn quando aquisição > 100ms.
    import time as _time
    _t0 = _time.monotonic()
    await db.execute(
        _text("SELECT pg_advisory_xact_lock(hashtextextended(:pid, 0))"),
        {"pid": str(project_id)},
    )
    _wait_ms = (_time.monotonic() - _t0) * 1000
    if _wait_ms > 100:
        logger.warning(
            "ingestion.advisory_lock_contention",
            wait_ms=round(_wait_ms, 1),
            project_id=str(project_id),
            lock_scope="dispatch_queue",
        )

    # Quantos já estão processando? Habilita paralelismo configurável.
    # Default=1 preserva FIFO sequencial; N>1 paraleliza até o teto.
    #
    # DECISÃO ARQUITETURAL F5.1 (Arq CR-1, DBA CR-1): só conta `processing`
    # como in_flight, NÃO `ocg_updating`. Doc em `ocg_updating` significa
    # pipeline n8n já terminou; OCGUpdater está rodando em Celery task.
    # Concorrência protegida pelo `pg_advisory_xact_lock` em
    # OCGUpdaterService.update_ocg_from_arguider — duas tasks Celery do
    # mesmo projeto serializam no advisory lock. Decisão intencional:
    # mais throughput (próximo doc inicia n8n enquanto OCG do anterior
    # consolida), com integridade garantida pelo lock cross-process.
    # NÃO "corrigir" pra incluir `ocg_updating` — sequencializaria sem ganho.
    max_parallel = getattr(settings, "INGESTION_MAX_PARALLEL_PER_PROJECT", 1)
    in_flight = await db.scalar(
        select(func.count(_IngDoc.id)).where(
            and_(
                _IngDoc.project_id == project_id,
                _IngDoc.arguider_status == "processing",
                _IngDoc.deleted_at.is_(None),
            )
        )
    )
    if in_flight and in_flight >= max_parallel:
        return None

    # Próximo pending mais antigo (exclui questionnaire — sintético, não pipeline).
    res = await db.execute(
        select(_IngDoc).where(
            and_(
                _IngDoc.project_id == project_id,
                _IngDoc.arguider_status == "pending",
                _IngDoc.deleted_at.is_(None),
                _IngDoc.file_type != "questionnaire",
            )
        ).order_by(_IngDoc.created_at.asc()).limit(1)
    )
    next_doc = res.scalar_one_or_none()
    if next_doc is None:
        return None

    # Marca processing antes de soltar o lock — outras requests verão in_flight=1.
    next_doc.arguider_status = "processing"
    next_doc.arguider_stage = "queued"
    await db.flush()
    doc_id = next_doc.id
    file_type = next_doc.file_type or ""

    # HITL fast-path canônico (antes do fork n8n/Celery).
    # Detecta marker `gca-followup-marker` em qualquer arquivo de texto e
    # processa as respostas direto, sem rodar pipeline. Se o marker está
    # presente mas o parse falha, doc vira erro com mensagem clara — NUNCA
    # cai pro pipeline (que avaliaria o frontmatter como conteúdo bizarro).
    from app.utils.ingested_storage import read_ingested
    file_bytes = read_ingested(project_id, next_doc.filename) or b""
    if file_bytes:
        try:
            text_preview = file_bytes.decode("utf-8", errors="ignore")
        except Exception:  # noqa: BLE001
            text_preview = ""
        if "gca-followup-marker" in text_preview:
            svc = IngestionService(db)
            try:
                handled = await svc._process_followup_upload(
                    doc_id=doc_id, project_id=project_id, text=text_preview,
                )
            except Exception as exc:  # noqa: BLE001
                logger.error(
                    "ingestion.hitl_marker_parse_failed",
                    document_id=str(doc_id),
                    error=str(exc),
                    exc_info=True,
                )
                handled = None

            if handled and handled.get("answered", 0) > 0:
                logger.info(
                    "ingestion.hitl_dispatched_from_queue",
                    document_id=str(doc_id),
                    project_id=str(project_id),
                    answered=handled["answered"],
                    persona_id=handled.get("persona_id"),
                )
                return doc_id

            # Falha ou 0 respostas: marcar erro com mensagem clara e NÃO
            # cair pro pipeline (que daria score=0 para todas as personas).
            err_msg = (
                "Arquivo HITL detectado (gca-followup-marker presente) mas "
                "o parse não extraiu respostas válidas. Verifique se o "
                "frontmatter tem `project_id` e `persona_id` corretos e se "
                "as respostas estão preenchidas no formato esperado "
                "(blocos `<!-- pfq-id: UUID -->` seguidos de `**Resposta**:`)."
            )
            next_doc.arguider_status = "error"
            next_doc.arguider_stage = "failed"
            next_doc.arguider_error_message = err_msg
            await db.commit()
            return doc_id

    # Bytes do storage (n8n precisa do payload base64; Celery faz por id).
    use_n8n = getattr(settings, "INGESTION_VIA_N8N", False)
    if use_n8n:
        await db.commit()
        svc = IngestionService(db)
        await svc._dispatch_to_n8n(doc_id, project_id, file_type, file_bytes)
    else:
        await db.commit()
        pipeline_ingest_task.delay(str(doc_id), str(project_id), file_type)

    logger.info(
        "ingestion.dispatched_from_queue",
        document_id=str(doc_id),
        project_id=str(project_id),
        pipeline="n8n" if use_n8n else "celery",
    )
    return doc_id


async def _reevaluate_gatekeeper_for_audit(
    db: AsyncSession,
    project_id: UUID,
    ocg_version: Optional[int],
    trigger: str,
) -> None:
    """Recomputa Gatekeeper com o OCG atual e grava evento no audit_log.

    Gatekeeper é consolidação read-only em tempo real: aplica thresholds no
    momento da leitura usando o OCG mais recente. Esta função força essa
    recomputação após uma mudança relevante de OCG (contrato §5) e deixa
    trilha auditável do estado resultante.

    - Se o projeto não tem OCG, é no-op silencioso (nada a reavaliar).
    - Não muta `GatekeeperItem` nem `ModuleCandidate` (permanecem como os
      Arguiders anteriores os gravaram).
    - Idempotente do ponto de vista de dados; cada chamada gera um novo
      evento de audit (trilha cronológica).
    """
    from app.services.gatekeeper_service import GatekeeperService
    from app.services.audit_service import AuditService

    ocg_check = await db.execute(
        select(OCG).where(OCG.project_id == project_id).order_by(OCG.created_at.desc()).limit(1)
    )
    if ocg_check.scalar_one_or_none() is None:
        logger.info(
            "ingestion.gatekeeper_reeval_noop_no_ocg",
            project_id=str(project_id),
            trigger=trigger,
        )
        return

    gatekeeper = GatekeeperService(db)
    state = await gatekeeper.get_project_gatekeeper(project_id)
    ocg_section = state.get("ocg", {}) or {}
    summary = state.get("summary", {}) or {}

    audit = AuditService(db)
    await audit.log_event(
        event_type="GATEKEEPER_REEVALUATED",
        resource_type="project",
        resource_id=project_id,
        details={
            "trigger": trigger,
            "ocg_version": ocg_version,
            "blocking_pillars": ocg_section.get("blocking_pillars", []),
            "derived_status": ocg_section.get("derived_status"),
            "overall_score": (ocg_section.get("status") or {}).get("overall_score"),
            "has_blockers": summary.get("has_blockers"),
            "open_gaps": summary.get("open_gaps"),
            "open_show_stoppers": summary.get("open_show_stoppers"),
        },
    )


async def _reevaluate_gatekeeper_async(
    project_id: UUID,
    ocg_version: Optional[int],
    trigger: str = "document_ingestion",
) -> None:
    """Wrapper fire-and-forget de `_reevaluate_gatekeeper_for_audit`.

    Abre sessão própria, commita o evento de audit. Isolado: se falhar, não
    afeta a transação de OCG/ingestão (já commitadas).
    """
    try:
        from app.db.database import AsyncSessionLocal

        async with AsyncSessionLocal() as db:
            await _reevaluate_gatekeeper_for_audit(
                db, project_id, ocg_version=ocg_version, trigger=trigger,
            )
            await db.commit()
    except Exception as exc:  # noqa: BLE001
        import traceback
        logger.warning(
            "ingestion.gatekeeper_reeval_failed",
            project_id=str(project_id),
            error=str(exc) or repr(exc),
            error_type=type(exc).__name__,
            traceback=traceback.format_exc(),
        )


async def _regenerate_backlog_for_audit(
    db: AsyncSession,
    project_id: UUID,
    ocg_version: Optional[int],
    trigger: str,
) -> None:
    """Regenera backlog do OCG atual e grava evento BACKLOG_REGENERATED.

    Usado em dois cenários sem `changes` estruturados:
    - **Geração inicial** do OCG a partir do questionário aprovado (não há
      diff; o backlog precisa ser populado do zero).
    - **Contração** no delete de documento (itens de versões obsoletas
      precisam ser substituídos pelo estado atual do OCG).

    Preserva items com `source="manual"`. No-op silencioso se o projeto não
    tem OCG. Não comita — caller decide.
    """
    from app.services.backlog_service import BacklogService
    from app.services.audit_service import AuditService

    ocg_check = await db.execute(
        select(OCG).where(OCG.project_id == project_id).order_by(OCG.created_at.desc()).limit(1)
    )
    if ocg_check.scalar_one_or_none() is None:
        logger.info(
            "ingestion.backlog_regen_noop_no_ocg",
            project_id=str(project_id),
            trigger=trigger,
        )
        return

    backlog_svc = BacklogService(db)
    result = await backlog_svc.regenerate_from_ocg(project_id, ocg_version)

    audit = AuditService(db)
    await audit.log_event(
        event_type="BACKLOG_REGENERATED",
        resource_type="backlog",
        resource_id=project_id,
        details={
            "trigger": trigger,
            "ocg_version": ocg_version,
            "regenerated": result.get("regenerated", 0),
        },
    )


async def _regenerate_backlog_async(
    project_id: UUID,
    ocg_version: Optional[int],
    trigger: str,
) -> None:
    """Wrapper fire-and-forget de `_regenerate_backlog_for_audit`.

    Abre sessão própria, commita o resultado. Isolado: se falhar, não afeta
    a transação de OCG (já commitada).
    """
    try:
        from app.db.database import AsyncSessionLocal

        async with AsyncSessionLocal() as db:
            await _regenerate_backlog_for_audit(
                db, project_id, ocg_version=ocg_version, trigger=trigger,
            )
            await db.commit()
    except Exception as exc:  # noqa: BLE001
        import traceback
        logger.warning(
            "ingestion.backlog_regen_failed",
            project_id=str(project_id),
            error=str(exc) or repr(exc),
            error_type=type(exc).__name__,
            traceback=traceback.format_exc(),
        )


async def _fire_ocg_change_hooks(
    project_id: UUID,
    ocg_version: Optional[int],
    trigger: str,
    changes: Optional[list[dict]] = None,
) -> None:
    """Dispara hooks de consistência downstream após uma mudança de OCG.

    Unifica os 3 pontos onde OCG pode mudar:
    - Ingestão de documento: passa `changes` do `OCGUpdaterService`
      (roteia via PropagationService para categorias afetadas).
    - Contração no delete: passa `changes` dos campos revertidos.
    - Geração inicial via questionário: passa `changes=None` (não há diff;
      regenera backlog do zero).

    Todos os hooks são fire-and-forget com sessão própria e erros isolados.
    """
    # MVP 13 Fase 13.3b: migrado para Celery. Bytes/state ficam no DB,
    # apenas IDs passam pelo broker. Retry bounded + ACK late.
    from app.tasks.pipeline import (
        propagate_task,
        regenerate_backlog_task,
        reevaluate_gatekeeper_task,
    )
    if changes:
        propagate_task.delay(str(project_id), list(changes), ocg_version)
    else:
        regenerate_backlog_task.delay(str(project_id), ocg_version, trigger)

    reevaluate_gatekeeper_task.delay(str(project_id), ocg_version, trigger)


class IngestionService:
    """Serviço de ingestão de documentos por projeto."""

    def __init__(self, db: AsyncSession):
        self.db = db

    async def _process_followup_upload(
        self,
        doc_id: UUID,
        project_id: UUID,
        text: str,
    ) -> Optional[dict]:
        """Processa upload com marcador `gca-followup-marker` (HITL offline).

        Parse do .md gerado por GET /pipeline-questions/personas/{p}/download:
          - Frontmatter YAML simples (gca-followup-marker, persona_id, etc).
          - Blocos com `<!-- pfq-id: UUID -->` + parágrafo "Resposta:".
          - Marca PFQs como `answered`, atualiza ingested_document para
            `file_type='persona_followup'` + completed, sem rodar n8n.

        Retorna dict com {answered, persona_id, total} ou None se parse falhar.
        """
        import re
        from sqlalchemy import select as _sel, update as _upd
        from app.models.base import IngestedDocument, PersonaFollowUpQuestion

        # 1. Validar marcador + extrair persona_id (frontmatter YAML)
        m_proj = re.search(r"project_id:\s*([0-9a-f-]{36})", text, re.IGNORECASE)
        m_pers = re.search(r"persona_id:\s*([A-Z]{2,5})", text)
        if not m_proj or not m_pers:
            return None
        if m_proj.group(1).lower() != str(project_id).lower():
            logger.warning(
                "ingestion.followup_marker_project_mismatch",
                doc_id=str(doc_id),
                marker_project=m_proj.group(1),
                actual_project=str(project_id),
            )
            return None
        persona_id_norm = m_pers.group(1).upper()

        # 2. Parse blocos `<!-- pfq-id: UUID -->` ... "Resposta:" ... próximo bloco/EOF
        # Pega o conteúdo a partir do comentário até o próximo "## Q" ou EOF.
        block_re = re.compile(
            r"<!--\s*pfq-id:\s*([0-9a-f-]{36})\s*-->(.*?)(?=<!--\s*pfq-id:|\Z)",
            re.IGNORECASE | re.DOTALL,
        )
        answer_re = re.compile(
            r"\*\*Resposta\*\*:\s*\n+(.+?)(?=\n##\s|$)",
            re.IGNORECASE | re.DOTALL,
        )
        parsed: dict[str, str] = {}
        for bm in block_re.finditer(text):
            pfq_id = bm.group(1)
            block = bm.group(2)
            am = answer_re.search(block)
            if not am:
                continue
            answer = am.group(1).strip()
            # Ignora placeholder não preenchido
            if not answer or answer.lower().startswith("<sua resposta") or answer == "<sua resposta aqui>":
                continue
            parsed[pfq_id] = answer

        if not parsed:
            return {"answered": 0, "persona_id": persona_id_norm, "total": 0}

        # 3. UPDATE PFQs (filtrando por persona+projeto pra evitar abuse)
        now = datetime.now(timezone.utc)
        updated = 0
        for pfq_id_str, answer in parsed.items():
            try:
                pfq_uuid = UUID(pfq_id_str)
            except (ValueError, TypeError):
                continue
            res = await self.db.execute(
                _upd(PersonaFollowUpQuestion)
                .where(
                    PersonaFollowUpQuestion.id == pfq_uuid,
                    PersonaFollowUpQuestion.project_id == project_id,
                    func.upper(PersonaFollowUpQuestion.persona_id) == persona_id_norm,
                    PersonaFollowUpQuestion.status == "pending",
                )
                .values(
                    answer_text=answer[:5000],
                    answer_provided_at=now,
                    status="answered",
                    updated_at=now,
                )
            )
            if (res.rowcount or 0) > 0:
                updated += 1

        # 4. Atualiza IngestedDocument: vira sintético persona_followup, completed
        await self.db.execute(
            _upd(IngestedDocument)
            .where(IngestedDocument.id == doc_id)
            .values(
                file_type="persona_followup",
                arguider_status="completed",
                arguider_stage="followup_synthetic",
                arguider_progress_percent=100,
                arguider_error_message=None,
                updated_at=now,
            )
        )
        await self.db.commit()

        # 5. F2 — Recompute do OCG considerando as respostas HITL.
        # Decisão arquitetural (gate DBA aprovado): chamada feita em
        # transação SEPARADA (nova sessão), após commit das PFQs, pra
        # respeitar a invariante de ordem de locks documentada em
        # ocg_updater_service.py:273 (asyncio.Lock antes do advisory).
        # Trade-off LLM vs heurística: optamos por LLM completo (~R$0,02
        # por HITL) — é robusto, sem refactor do OCGUpdaterService, e o
        # próprio LLM julga se as respostas justificam expansão (preserva
        # o invariante "OCG só expande quando recebe informação de valor"
        # do CLAUDE.md GCA §2.4).
        if updated > 0:
            try:
                from app.db.database import AsyncSessionLocal
                from app.services.ocg_updater_service import (
                    OCGUpdaterService,
                    TRIGGER_HITL_FOLLOWUP,
                )

                # F2 FIX: shape sintético DENTRO do contrato esperado pelo
                # OCGUpdaterService — chaves ocg_individual/ocg_global_delta/
                # recommendations. Sem isso o LLM consolidador não sabia
                # interpretar `hitl_answers` e retornava delta vazio.
                # Buscar texto das perguntas (não só pfq_id UUID) pra prompt.
                from sqlalchemy import select as _select
                from app.models.base import PersonaFollowUpQuestion as _PFQ
                pfq_rows = (await self.db.execute(
                    _select(_PFQ.id, _PFQ.question_text, _PFQ.persona_name)
                    .where(_PFQ.id.in_([UUID(pid) for pid in parsed.keys()]))
                )).all()
                pfq_text_map = {str(r[0]): (r[1], r[2]) for r in pfq_rows}

                # Monta lista pergunta-resposta como `recommendations` —
                # campo que o updater consolidador interpreta como input rico.
                hitl_qa = []
                for pfq_id, ans in parsed.items():
                    qtext, pname = pfq_text_map.get(
                        pfq_id, ("(pergunta perdida)", persona_id_norm)
                    )
                    hitl_qa.append({
                        "pfq_id": pfq_id,
                        "persona": persona_id_norm,
                        "persona_name": pname or persona_id_norm,
                        "pergunta": qtext or "(texto não encontrado)",
                        "resposta": ans,
                        "tipo": "hitl_followup_answer",
                    })

                # Sintético no shape canônico — HITL é "fonte de verdade
                # adicional" pra persona específica. Inclui o parecer da
                # persona se já existe (preserva MAX-por-persona).
                from app.models.base import OCGIndividual as _OI
                last_persona = (await self.db.execute(
                    _select(_OI.parecer)
                    .where(
                        _OI.project_id == project_id,
                        _OI.persona_id == persona_id_norm,
                        _OI.status == "completed",
                    )
                    .order_by(_OI.completed_at.desc())
                    .limit(1)
                )).scalar_one_or_none()
                persona_parecer = last_persona or {}

                arguider_analysis_synthetic = {
                    "trigger_source": "hitl_followup",
                    "persona_id": persona_id_norm,
                    "answered_count": updated,
                    # Chaves canônicas que o OCGUpdater espera:
                    "ocg_individual": {persona_id_norm: persona_parecer},
                    "ocg_global_delta": {},
                    "gaps": [],
                    "show_stoppers": [],
                    "recommendations": hitl_qa,
                    # Hint extra pra LLM consolidador: aqui temos respostas
                    # diretas do GP a perguntas que A PRÓPRIA persona havia
                    # gerado (filosofia Assistida). É informação de alta
                    # confiança — deve elevar o pilar correspondente.
                    "_hitl_hint": (
                        f"GP respondeu {updated} pergunta(s) em aberto da "
                        f"persona {persona_id_norm}. Cada par "
                        f"pergunta/resposta em 'recommendations' é resposta "
                        f"direta do owner — preencheu lacunas que a persona "
                        f"havia identificado. Usar pra elevar score do "
                        f"pilar correspondente, não descartar."
                    ),
                }

                async with AsyncSessionLocal() as ocg_db:
                    ocg_svc = OCGUpdaterService(ocg_db)
                    ocg_result = await ocg_svc.update_ocg_from_arguider(
                        project_id=project_id,
                        arguider_analysis=arguider_analysis_synthetic,
                        document_id=doc_id,
                        actor_id=None,
                        trigger_source=TRIGGER_HITL_FOLLOWUP,
                    )
                    logger.info(
                        "ingestion.hitl_ocg_updated",
                        document_id=str(doc_id),
                        project_id=str(project_id),
                        persona_id=persona_id_norm,
                        answered=updated,
                        version_to=(ocg_result or {}).get("version_to"),
                    )

                # 5b. Audit dedicado HITL (DBA C4) — registra pfq_ids
                # respondidos pra rastreabilidade. correlation_id aponta
                # ao IngestedDocument HITL pra reconstituir a cadeia.
                async with AsyncSessionLocal() as audit_db:
                    from app.services.audit_service import AuditService
                    aud = AuditService(audit_db)
                    await aud.log_event(
                        event_type="HITL_RESPONSES_RECEIVED",
                        resource_type="ingested_document",
                        resource_id=doc_id,
                        details={
                            "pfq_ids": list(parsed.keys()),
                            "persona_id": persona_id_norm,
                            "answered_count": updated,
                            "total_in_upload": len(parsed),
                        },
                        correlation_id=doc_id,
                    )
                    await audit_db.commit()
            except Exception as exc:  # noqa: BLE001
                # Falha do recompute não derruba o registro HITL — as PFQs
                # já estão persistidas como answered. Logar pra investigação.
                logger.error(
                    "ingestion.hitl_ocg_recompute_failed",
                    document_id=str(doc_id),
                    project_id=str(project_id),
                    error=str(exc),
                    exc_info=True,
                )

        # 6. Dispara próximo da fila (este doc não vai pro pipeline)
        try:
            await dispatch_first_pending_for_project(self.db, project_id)
        except Exception as exc:  # noqa: BLE001
            logger.warning("ingestion.followup_dispatch_next_failed", error=str(exc))

        return {
            "answered": updated,
            "persona_id": persona_id_norm,
            "total": len(parsed),
        }

    async def _dispatch_to_n8n(self, doc_id, project_id, file_type, file_bytes):
        """Dispara o pipeline de ingestão via webhook n8n (feature flag INGESTION_VIA_N8N).

        MVP 35 (Arq-M1): file_type='questionnaire' é IngestedDocument sintético
        criado direto no router de submit do questionário — NÃO passa pelo n8n.
        Guard explícito aqui evita dispatch acidental se outra rota chamar
        `upload_document` com esse tipo.
        """
        if file_type == "questionnaire":
            logger.info(
                "ingestion.dispatch_skipped_questionnaire_synthetic",
                document_id=str(doc_id),
                project_id=str(project_id),
            )
            return
        import base64
        import json as _json
        import hmac
        import hashlib
        import httpx
        from datetime import datetime, timezone
        from app.core.config import settings
        from app.services.ai_key_resolver import AIKeyResolver
        from app.services.personas.prompts_registry import PERSONA_PROMPTS

        provider_chain = await AIKeyResolver.resolve_project_provider_chain(
            self.db, project_id, include_api_key=True
        )
        chain_data = [
            {"provider": p.get("provider", ""), "model": p.get("model", ""), "api_key": p.get("api_key", "")}
            for p in (provider_chain or [])
        ]

        # Pegar metadata do doc recém-criado
        from app.models.base import IngestedDocument
        doc = await self.db.get(IngestedDocument, doc_id)

        # F1-C2 (Arquiteto): Para imagens, filtra a chain por providers
        # com capability Vision. DeepSeek/Ollama-text recebendo payload
        # multimodal falha silenciosamente — pré-filtra no backend.
        # Reutiliza PROVIDERS_WITH_VISION canônico de vision_service.
        if file_type == "image":
            from app.services.vision_service import PROVIDERS_WITH_VISION
            chain_data = [c for c in chain_data if c["provider"] in PROVIDERS_WITH_VISION]
            if not chain_data:
                # Sem provider Vision → notifica GPs e curto-circuita
                # (não enfileira no n8n — config gap, não erro de sistema).
                try:
                    from app.services.notification_inapp_service import InAppNotificationService
                    from app.models.base import ProjectMember
                    gps_q = await self.db.execute(
                        select(ProjectMember).where(
                            ProjectMember.project_id == project_id,
                            ProjectMember.role == "gp",
                            ProjectMember.is_active == True,
                        )
                    )
                    notif = InAppNotificationService(self.db)
                    for gp in gps_q.scalars().all():
                        await notif.notify(
                            user_id=gp.user_id,
                            event_type="vision_provider_missing",
                            title="Provider com Visão não configurado",
                            message=(
                                "A imagem enviada exige um provedor de IA com "
                                "capacidade de Visão (Anthropic ou OpenAI). "
                                "Configure em Configurações → IA antes de re-enviar."
                            ),
                            severity="warning",
                            project_id=project_id,
                            resource_type="ingested_document",
                            resource_id=doc_id,
                        )
                except Exception as exc:  # noqa: BLE001
                    logger.warning("ingestion.vision_notify_failed", error=str(exc))
                # Marca doc como erro com mensagem clara
                if doc:
                    doc.arguider_status = "error"
                    doc.arguider_stage = "failed"
                    doc.arguider_error_message = (
                        "Imagem requer provedor com Vision (Anthropic ou OpenAI). "
                        "Configure em Configurações → IA."
                    )
                    await self.db.commit()
                return

        # Mime type a partir do file_type
        mime_map = {
            "pdf": "application/pdf",
            "docx": "application/vnd.openxmlformats-officedocument.wordprocessingml.document",
            "markdown": "text/markdown",
            "text": "text/plain",
            "code": "text/plain",
            # B2 (2026-05-04): novos formatos. Pré-extração converte para
            # text/plain antes do dispatch — mas se a extração falhar e
            # cair pro fallback bytes-as-is, mime_type indica o original.
            "html": "text/html",
            "rtf": "application/rtf",
            "eml": "message/rfc822",
            "spreadsheet": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        }
        # F1-C1 (Arquiteto): Para image, deriva mime da extensão real do
        # arquivo. mime_map.get('image') retornaria octet-stream e a API
        # Anthropic rejeitaria o payload (mime inválido para source.type='base64').
        if file_type == "image" and doc:
            ext_to_mime = {
                "png": "image/png", "jpg": "image/jpeg", "jpeg": "image/jpeg",
                "gif": "image/gif", "webp": "image/webp",
            }
            ext = (doc.original_filename or "").rsplit(".", 1)[-1].lower()
            mime_type = ext_to_mime.get(ext, "image/png")  # fallback seguro
            if ext == "gif":
                logger.info("ingestion.gif_first_frame_only", doc_id=str(doc_id))

            # F1-C5 (Arquiteto): registra evento de extração Vision com
            # estimativa de tokens. Cliente AJA precisa de rastreabilidade
            # de custo por documento (50 prints WhatsApp ≈ 80k tokens).
            # Estimativa: ~1600 tokens input por imagem 1024px (Anthropic),
            # ~200 tokens output (texto extraído). Refinamento real chega
            # via callback /ingestion-complete e sobrescreve.
            try:
                from app.services.ai_billing_service import AIBillingService
                if chain_data:  # já filtrado por PROVIDERS_WITH_VISION acima
                    billing = AIBillingService(self.db)
                    await billing.log_usage(
                        project_id=project_id,
                        provider=chain_data[0]["provider"],
                        model=chain_data[0].get("model") or "vision-default",
                        operation="vision_extraction_estimate",
                        tokens_input=1600,
                        tokens_output=200,
                        metadata={
                            "doc_id": str(doc_id),
                            "filename": doc.original_filename,
                            "mime_type": mime_type,
                            "estimate": True,
                        },
                    )
            except Exception as exc:  # noqa: BLE001
                logger.warning("ingestion.vision_billing_log_failed", error=str(exc))
        else:
            mime_type = mime_map.get(file_type, "application/octet-stream")

        # DOCX/PDF/HTML/RTF/EML/SPREADSHEET: Normalizer n8n não tem extrator
        # nativo (sem require()). Pré-extraímos texto aqui e enviamos como
        # text/plain. B2 (auditoria 2026-05-04) adicionou html/rtf/eml/xlsx.
        # Política de erro: try/except com log estruturado (Decisão 8).
        payload_bytes = file_bytes
        if file_type == "html" and file_bytes:
            try:
                import html2text
                h = html2text.HTML2Text()
                h.ignore_images = True
                h.body_width = 0  # sem wrap
                html_text = h.handle(file_bytes.decode("utf-8", errors="replace"))
                if html_text.strip():
                    payload_bytes = html_text.encode("utf-8")
                    mime_type = "text/plain"
                else:
                    logger.warning("ingestion.html_extract_empty", doc_id=str(doc_id))
            except Exception as exc:  # noqa: BLE001
                logger.warning(
                    "ingestion.html_extract_failed",
                    doc_id=str(doc_id), error=str(exc),
                )
        elif file_type == "rtf" and file_bytes:
            try:
                from striprtf.striprtf import rtf_to_text
                rtf_text = rtf_to_text(file_bytes.decode("utf-8", errors="replace"))
                if rtf_text.strip():
                    payload_bytes = rtf_text.encode("utf-8")
                    mime_type = "text/plain"
                else:
                    logger.warning("ingestion.rtf_extract_empty", doc_id=str(doc_id))
            except Exception as exc:  # noqa: BLE001
                logger.warning(
                    "ingestion.rtf_extract_failed",
                    doc_id=str(doc_id), error=str(exc),
                )
        elif file_type == "eml" and file_bytes:
            try:
                from email import policy
                from email.parser import BytesParser
                msg = BytesParser(policy=policy.default).parsebytes(file_bytes)
                # Header + body texto. Anexos ignorados (vão como upload separado).
                lines = [
                    f"De: {msg.get('From','')}",
                    f"Para: {msg.get('To','')}",
                    f"Assunto: {msg.get('Subject','')}",
                    f"Data: {msg.get('Date','')}",
                    "",
                ]
                body = msg.get_body(preferencelist=("plain", "html"))
                if body is not None:
                    body_text = body.get_content() or ""
                    if body.get_content_type() == "text/html":
                        try:
                            import html2text as _h2t
                            h = _h2t.HTML2Text(); h.ignore_images = True; h.body_width = 0
                            body_text = h.handle(body_text)
                        except Exception:  # noqa: BLE001
                            pass
                    lines.append(body_text)
                eml_text = "\n".join(lines)
                if eml_text.strip():
                    payload_bytes = eml_text.encode("utf-8")
                    mime_type = "text/plain"
                else:
                    logger.warning("ingestion.eml_extract_empty", doc_id=str(doc_id))
            except Exception as exc:  # noqa: BLE001
                logger.warning(
                    "ingestion.eml_extract_failed",
                    doc_id=str(doc_id), error=str(exc),
                )
        elif file_type == "spreadsheet" and file_bytes:
            try:
                import openpyxl
                from io import BytesIO
                wb = openpyxl.load_workbook(BytesIO(file_bytes), data_only=True, read_only=True)
                lines = []
                for sheet_name in wb.sheetnames:
                    sh = wb[sheet_name]
                    lines.append(f"\n=== Planilha: {sheet_name} ===\n")
                    for row in sh.iter_rows(values_only=True):
                        # Pula linhas totalmente vazias
                        if any(c is not None and str(c).strip() for c in row):
                            lines.append(" | ".join(str(c) if c is not None else "" for c in row))
                xlsx_text = "\n".join(lines)
                if xlsx_text.strip():
                    payload_bytes = xlsx_text.encode("utf-8")
                    mime_type = "text/plain"
                else:
                    logger.warning("ingestion.xlsx_extract_empty", doc_id=str(doc_id))
            except Exception as exc:  # noqa: BLE001
                logger.warning(
                    "ingestion.xlsx_extract_failed",
                    doc_id=str(doc_id), error=str(exc),
                )
        elif file_type == "docx" and file_bytes:
            try:
                from app.services.rich_docx_extractor import extract_rich_text
                docx_text = extract_rich_text(file_bytes) or ""
                if docx_text.strip() and not docx_text.startswith("["):
                    payload_bytes = docx_text.encode("utf-8")
                    mime_type = "text/plain"
                else:
                    logger.warning(
                        "ingestion.docx_extract_empty",
                        doc_id=str(doc_id),
                        sample=docx_text[:120],
                    )
            except Exception as exc:  # noqa: BLE001
                logger.warning(
                    "ingestion.docx_extract_failed",
                    doc_id=str(doc_id),
                    error=str(exc),
                )
        elif file_type == "pdf" and file_bytes:
            # Pipeline OCR canônico em 2 camadas (LLM-agnóstico):
            #   1) pdfplumber: texto nativo (PDFs com camada texto, ~95%).
            #   2) Tesseract: pdf2image renderiza página → pytesseract por
            #      página onde a Camada 1 retornou pouco texto. Funciona
            #      em scans, apresentações com texto raster, screenshots.
            # Camada 3 (Vision LLM) fica para MVP futuro — depende de
            # provider opcional configurado pelo cliente em Settings → IA.
            MIN_CHARS_PER_PAGE = 50  # abaixo disso, considera "imagem/scan"
            try:
                import pdfplumber
                pages_native: list[str] = []
                pages_to_ocr: list[int] = []
                with pdfplumber.open(io.BytesIO(file_bytes)) as pdf:
                    for idx, page in enumerate(pdf.pages):
                        try:
                            t = (page.extract_text() or "").strip()
                        except Exception:  # noqa: BLE001
                            t = ""
                        if len(t) >= MIN_CHARS_PER_PAGE:
                            pages_native.append(t)
                        else:
                            pages_native.append("")  # placeholder pra ordem
                            pages_to_ocr.append(idx)

                ocr_pages_done = 0
                if pages_to_ocr:
                    try:
                        from pdf2image import convert_from_bytes
                        import pytesseract
                        # DPI 200 = bom equilíbrio qualidade/velocidade.
                        # Renderiza só as páginas que precisam.
                        for page_idx in pages_to_ocr:
                            try:
                                imgs = convert_from_bytes(
                                    file_bytes,
                                    dpi=200,
                                    first_page=page_idx + 1,
                                    last_page=page_idx + 1,
                                )
                                if not imgs:
                                    continue
                                ocr_text = pytesseract.image_to_string(
                                    imgs[0], lang="por+eng"
                                ).strip()
                                if ocr_text:
                                    pages_native[page_idx] = ocr_text
                                    ocr_pages_done += 1
                            except Exception as exc:  # noqa: BLE001
                                logger.warning(
                                    "ingestion.pdf_ocr_page_failed",
                                    doc_id=str(doc_id),
                                    page=page_idx,
                                    error=str(exc),
                                )
                    except ImportError:
                        logger.warning(
                            "ingestion.pdf_tesseract_unavailable",
                            doc_id=str(doc_id),
                            detail="pdf2image/pytesseract não instalados (requer rebuild da imagem)",
                        )

                pdf_text = "\n\n".join(p for p in pages_native if p).strip()
                if pdf_text:
                    payload_bytes = pdf_text.encode("utf-8")
                    mime_type = "text/plain"
                    logger.info(
                        "ingestion.pdf_extracted",
                        doc_id=str(doc_id),
                        chars=len(pdf_text),
                        pages_native=sum(1 for p in pages_native if p) - ocr_pages_done,
                        pages_ocr=ocr_pages_done,
                        pages_total=len(pages_native),
                    )
                else:
                    logger.warning(
                        "ingestion.pdf_no_text",
                        doc_id=str(doc_id),
                        detail="Nem pdfplumber nem Tesseract extraíram texto. Documento provavelmente é só imagem sem texto reconhecível.",
                    )
            except Exception as exc:  # noqa: BLE001
                logger.warning(
                    "ingestion.pdf_extract_failed",
                    doc_id=str(doc_id),
                    error=str(exc),
                )

        # HITL fast-path: se o texto extraído tem marcador "gca-followup-marker",
        # GP fez upload de respostas a perguntas em aberto. Processa direto no
        # backend (sem n8n personas): marca PFQs como answered + cria
        # IngestedDocument file_type='persona_followup'. Atalho determinístico.
        try:
            text_for_marker = (payload_bytes or b"").decode("utf-8", errors="ignore") if payload_bytes else ""
        except Exception:  # noqa: BLE001
            text_for_marker = ""
        if "gca-followup-marker" in text_for_marker:
            try:
                handled = await self._process_followup_upload(
                    doc_id=doc_id,
                    project_id=project_id,
                    text=text_for_marker,
                )
                if handled:
                    logger.info(
                        "ingestion.followup_marker_processed",
                        doc_id=str(doc_id),
                        project_id=str(project_id),
                        answered=handled.get("answered", 0),
                    )
                    return  # Não envia pra n8n
            except Exception as exc:  # noqa: BLE001
                logger.warning(
                    "ingestion.followup_marker_failed",
                    doc_id=str(doc_id),
                    error=str(exc),
                )
                # Fallthrough: trata como doc normal se parsing falhar

        # Seed do shared_context: respostas do Questionário Técnico aprovado +
        # snapshot do OCG atual. Sem isto cada persona analisa o doc no vácuo
        # (DBA não sabe que é SQL, SEG não sabe que é OAuth2 etc). Conferente
        # funde com seu summary e propaga aos specialists no dispatch.
        #
        # B4 (Decisão GP 2026-05-05): `ocg_summary` injetado é o respaldo do
        # princípio anti-retrabalho. O prompt das 12 personas instrui a
        # verificar shared_context.ocg_summary ANTES de gerar question[].
        # Sem este campo, instrução cai no vazio.
        from app.models.base import TechnicalQuestionnaire as _TQ
        from app.models.base import OCG as _OCG
        tq_row = (
            await self.db.execute(
                select(_TQ)
                .where(
                    _TQ.project_id == project_id,
                    _TQ.status == "submitted",
                )
                .order_by(_TQ.submitted_at.desc().nullslast())
                .limit(1)
            )
        ).scalar_one_or_none()

        # OCG atual — campo top-level + campos canônicos que personas consultam
        ocg_row = (
            await self.db.execute(
                select(_OCG)
                .where(_OCG.project_id == project_id)
                .order_by(_OCG.created_at.desc())
                .limit(1)
            )
        ).scalar_one_or_none()
        ocg_summary: dict = {}
        if ocg_row is not None and ocg_row.ocg_data:
            try:
                import json as _json
                ocg_full = _json.loads(ocg_row.ocg_data) if isinstance(ocg_row.ocg_data, str) else ocg_row.ocg_data
                ocg_summary = {
                    "version": ocg_row.version,
                    "overall_score": ocg_row.overall_score,
                    "status": ocg_row.status,
                    # Campos canônicos que personas verificam antes de perguntar:
                    "PROJECT_PROFILE": ocg_full.get("PROJECT_PROFILE") or {},
                    "PILLAR_SCORES": ocg_full.get("PILLAR_SCORES") or {},
                    "STACK_RECOMMENDATION": ocg_full.get("STACK_RECOMMENDATION") or {},
                    "ARCHITECTURE_OVERVIEW": ocg_full.get("ARCHITECTURE_OVERVIEW") or {},
                    "COMPLIANCE_CHECKLIST": ocg_full.get("COMPLIANCE_CHECKLIST") or {},
                    "TESTING_REQUIREMENTS": ocg_full.get("TESTING_REQUIREMENTS") or {},
                    "DELIVERABLES": ocg_full.get("DELIVERABLES") or {},
                    "RISK_ANALYSIS": ocg_full.get("RISK_ANALYSIS") or {},
                    "LGPD_COMPLIANCE": ocg_full.get("LGPD_COMPLIANCE") or {},
                    "APPROVAL_STATUS": ocg_full.get("APPROVAL_STATUS"),
                }
            except Exception as exc:  # noqa: BLE001
                logger.warning(
                    "ingestion.ocg_summary_parse_failed",
                    project_id=str(project_id), error=str(exc),
                )

        seed_shared_context: dict = {}
        if tq_row is not None:
            seed_shared_context["questionnaire_responses"] = tq_row.responses or {}
            seed_shared_context["questionnaire_submitted_at"] = (
                tq_row.submitted_at.isoformat() if tq_row.submitted_at else None
            )
        if ocg_summary:
            seed_shared_context["ocg_summary"] = ocg_summary

        n8n_payload = {
            "ingestion_id": str(doc_id),
            "project_id": str(project_id),
            "document_bytes_base64": base64.b64encode(payload_bytes).decode() if payload_bytes else "",
            "document_metadata": {
                "filename": (doc.original_filename if doc else "") or (doc.filename if doc else ""),
                "mime_type": mime_type,
                "size_bytes": len(payload_bytes) if payload_bytes else 0,
                "uploaded_by": str(doc.uploaded_by) if doc else "",
                "uploaded_by_role": "GP",
                "declared_purpose": "upload",
            },
            "provider_chain": chain_data,
            "persona_prompts": dict(PERSONA_PROMPTS),
            "seed_shared_context": seed_shared_context,
            "callback_url": f"{getattr(settings, 'GCA_CALLBACK_BASE_URL', 'http://gca-backend:8000')}/api/v1/webhooks/ingestion-complete",
            "timestamp": datetime.now(timezone.utc).isoformat(),
        }

        n8n_url = getattr(settings, "N8N_BASE_URL", "http://n8n:5678")
        webhook_secret = getattr(settings, "GCA_WEBHOOK_SECRET", "")
        body_bytes = _json.dumps(n8n_payload).encode()
        sig = "sha256=" + hmac.new(webhook_secret.encode(), body_bytes, hashlib.sha256).hexdigest()

        # Marcar status processing
        doc.arguider_status = "processing"
        doc.arguider_stage = "n8n_pipeline"
        await self.db.commit()

        async with httpx.AsyncClient(timeout=10) as client:
            resp = await client.post(
                f"{n8n_url}/webhook/gca-normalizer",
                content=body_bytes,
                headers={"Content-Type": "application/json", "X-GCA-Signature": sig},
            )
            if resp.status_code not in (200, 202):
                logger.error("ingestion.n8n_dispatch_failed", status=resp.status_code, body=resp.text[:200])
                raise RuntimeError(f"n8n dispatch falhou: {resp.status_code}")

    async def upload_document(
        self,
        project_id: UUID,
        uploaded_by: UUID,
        file_bytes: bytes,
        original_filename: str,
        content_type: str = "",
        category: str | None = None,
        target_module_id: UUID | None = None,
    ) -> dict:
        """Upload e análise assíncrona de documento.

        MVP 9 Fase 9.5.2 — `target_module_id` opcional vincula o doc a
        um item do Roadmap. Quando o pipeline confirma propagação, o
        item transita pra `adicionado` e cria DELIVERABLE automático.
        Se `target_module_id` for None E o file_type for pdf, tenta
        extrair de hidden field/metadata/footer do template GCA.
        """
        # Validar tamanho
        if len(file_bytes) > MAX_FILE_SIZE:
            return {"error": "Arquivo excede o tamanho máximo de 50MB", "status_code": 413}

        # Determinar tipo
        ext = original_filename.rsplit(".", 1)[-1].lower() if "." in original_filename else ""
        file_type = EXTENSION_MAP.get(ext)
        if not file_type:
            return {"error": f"Tipo de arquivo '.{ext}' não suportado", "status_code": 400}

        # SHA256 para deduplicação
        file_hash = hashlib.sha256(file_bytes).hexdigest()

        # Verificar duplicata
        # MVP 34 (DBA-M1): filtra docs soft-deleted — caso contrário, re-ingestão
        # de versão anonimizada de doc deletado por LGPD seria bloqueada.
        existing = await self.db.execute(
            select(IngestedDocument).where(
                IngestedDocument.project_id == project_id,
                IngestedDocument.file_hash == file_hash,
                IngestedDocument.deleted_at.is_(None),
            )
        )
        dup = existing.scalar_one_or_none()
        if dup:
            return {
                "duplicate": True,
                "existing_document_id": str(dup.id),
                "message": "Documento já ingerido neste projeto.",
                "status_code": 409,
            }

        # Gerar filename único
        filename = f"{uuid4()}.{ext}"

        # Persistir bytes em storage para abertura read-only posterior
        from app.utils.ingested_storage import write_ingested
        write_ingested(project_id, filename, file_bytes)

        # PII detection — triagem básica no conteúdo textual
        pii_detected, pii_fields = self._detect_pii(file_bytes, file_type)
        quarantine_status = "quarantined" if pii_detected else "none"

        # MVP 9 Fase 9.5.2 — descobre target_module_id se ausente.
        # Estratégia: explícito do form > hidden field do PDF > metadata > footer.
        if target_module_id is None and file_type == "pdf":
            from app.services.pdf_module_id_extractor import extract_module_id
            extracted = extract_module_id(file_bytes)
            if extracted is not None:
                # Validação de compartimentalização (§2.2): módulo precisa
                # pertencer ao mesmo projeto. Se não, ignora silenciosamente
                # — não é falha, é doc upload normal.
                from app.models.base import ModuleCandidate as _MC
                mc_check = await self.db.get(_MC, extracted)
                if mc_check and mc_check.project_id == project_id:
                    target_module_id = extracted
                    logger.info(
                        "ingestion.target_module_extracted_from_pdf",
                        document_filename=original_filename,
                        target_module_id=str(target_module_id),
                    )
                else:
                    logger.warning(
                        "ingestion.target_module_pdf_cross_project",
                        extracted=str(extracted),
                        project_id=str(project_id),
                    )

        # Reforma Arguidor #1 (2026-04-25): auto-detecção de doc canônico.
        # Decisão soberana do owner — substitui valores antigos no OCG.
        canonical_markers = (
            "decisoes_canonicas", "canonical_decision", "ata_decisoes",
            "decision_record", "rfc_canonica",
        )
        haystack = (
            f"{(original_filename or '').lower()} {(category or '').lower()}"
        )
        is_canonical = any(m in haystack for m in canonical_markers)

        # Criar registro
        document = IngestedDocument(
            project_id=project_id,
            filename=filename,
            original_filename=original_filename,
            file_type=file_type,
            document_category=category,
            file_hash=file_hash,
            file_size_bytes=len(file_bytes),
            uploaded_by=uploaded_by,
            quarantine_status=quarantine_status,
            pii_detected=pii_detected,
            pii_fields=json.dumps(pii_fields) if pii_fields else None,
            arguider_status="pending" if not pii_detected else "quarantined",
            git_file_path=f"docs/ingested/uncategorized/{filename}",
            target_module_id=target_module_id,
            is_canonical_decision=is_canonical,
        )
        self.db.add(document)
        await self.db.commit()

        doc_id = document.id

        # M01 — auto-linkagem de resposta de iteração.
        # Se o documento carrega o marker `gca_iteration_id=<uuid>` nos
        # metadados (DOCX `core_properties.keywords` preferencial, PDF
        # `Keywords` legado), linka o doc à iteração correspondente
        # automaticamente. Ingestão vira ponto único de entrada — usuário
        # não precisa mais subir pela aba Questões em Aberto.
        if file_type in ("docx", "pdf"):
            try:
                iteration_id_marker = _extract_m01_iteration_marker(file_bytes, file_type)
                if iteration_id_marker is not None:
                    from app.models.base import CustomQuestionnaireIteration
                    iter_result = await self.db.execute(
                        select(CustomQuestionnaireIteration).where(
                            (CustomQuestionnaireIteration.id == iteration_id_marker)
                            & (CustomQuestionnaireIteration.project_id == project_id)
                            & (CustomQuestionnaireIteration.status == "pending")
                        )
                    )
                    iter_row = iter_result.scalar_one_or_none()
                    if iter_row is not None:
                        iter_row.answer_document_id = doc_id
                        await self.db.commit()
                        logger.info(
                            "ingestion.m01_auto_linked",
                            document_id=str(doc_id),
                            iteration_id=str(iteration_id_marker),
                            project_id=str(project_id),
                        )
            except Exception as exc:  # noqa: BLE001
                # Falha na auto-linkagem nunca deve bloquear a ingestão
                # canônica — só reporta.
                logger.warning(
                    "ingestion.m01_auto_link_failed",
                    document_id=str(doc_id),
                    error=str(exc),
                )

        if pii_detected:
            logger.warning(
                "ingestion.pii_detected_quarantined",
                document_id=str(doc_id),
                pii_fields=pii_fields,
            )
            return {
                "document_id": str(doc_id),
                "quarantined": True,
                "pii_fields": pii_fields,
                "message": "Documento em quarentena — PII detectado. Requer decisão explícita.",
                "status_code": 200,
            }

        # MVP 25 Fase 25.3 — Design tokens via Ingestão de stylesheet.
        # CSS/SCSS vai pro extractor determinístico (zero LLM) e grava no
        # OCG.STACK_RECOMMENDATION.frontend.design_tokens. Ignora pipeline
        # Celery padrão — não há análise Arguidor pra stylesheet.
        design_tokens_applied: Optional[dict] = None
        if file_type == "stylesheet":
            try:
                from app.services.css_token_extractor_service import extract_tokens
                from app.services.design_tokens import from_extractor_output
                from app.services.design_tokens_applier_service import (
                    apply_tokens_to_ocg,
                )

                css_text = file_bytes.decode("utf-8", errors="ignore")
                extracted = extract_tokens(css_text)
                if not extracted.is_empty:
                    # Se já existe tokens no OCG, passa previous pra marcar "mixed"
                    ocg_row = (await self.db.execute(
                        select(OCG).where(OCG.project_id == project_id)
                        .order_by(OCG.created_at.desc()).limit(1)
                    )).scalar_one_or_none()
                    previous = None
                    if ocg_row and ocg_row.ocg_data:
                        try:
                            data = json.loads(ocg_row.ocg_data)
                            previous = (
                                (data.get("STACK_RECOMMENDATION") or {})
                                .get("frontend") or {}
                            ).get("design_tokens")
                        except (TypeError, ValueError):
                            previous = None

                    payload = from_extractor_output(extracted, previous=previous)
                    result = await apply_tokens_to_ocg(
                        self.db, project_id, payload,
                        actor_id=uploaded_by,
                        source_document_id=doc_id,
                    )
                    document.document_category = "design_stylesheet"
                    document.arguider_status = "completed"
                    document.arguider_completed_at = datetime.now(timezone.utc)
                    await self.db.commit()
                    design_tokens_applied = {
                        "applied": result["applied"],
                        "ocg_version_to": result["ocg_version_to"],
                        "tokens_preview": {
                            "palette_top": list(extracted.palette_top[:6]),
                            "families": list(extracted.font_families[:3]),
                            "unique_colors": extracted.colors_unique_count,
                        },
                    }
                    logger.info(
                        "ingestion.design_tokens_extracted",
                        document_id=str(doc_id),
                        project_id=str(project_id),
                        **design_tokens_applied,
                    )
                else:
                    # CSS sem tokens detectáveis — marca como processado,
                    # não reclama (pode ser reset.css ou similar).
                    document.arguider_status = "completed"
                    document.arguider_completed_at = datetime.now(timezone.utc)
                    document.document_category = "design_stylesheet_empty"
                    await self.db.commit()
                    logger.info(
                        "ingestion.stylesheet_no_tokens",
                        document_id=str(doc_id),
                    )
            except Exception as exc:
                # Não bloqueia upload; deixa doc disponível pra inspeção manual.
                logger.warning(
                    "ingestion.design_tokens_hook_failed",
                    document_id=str(doc_id),
                    error=str(exc)[:300],
                )

        # Stylesheet sempre pula o pipeline LLM — com ou sem tokens detectados.
        if file_type == "stylesheet":
            if design_tokens_applied is not None:
                return {
                    "document_id": str(doc_id),
                    "status": "completed",
                    "design_tokens_applied": design_tokens_applied,
                    "message": (
                        f"Design tokens extraídos. OCG "
                        f"v{design_tokens_applied['ocg_version_to']}."
                    ),
                }
            return {
                "document_id": str(doc_id),
                "status": "completed",
                "message": "Stylesheet ingerido sem tokens detectáveis.",
            }

        # Para mocks visuais (PNG/PDF) ingeridos sem design tokens ainda no
        # OCG, abre gap canônico no Arguidor pedindo paleta/tipografia.
        # Best-effort — não bloqueia upload se o seed falhar.
        if file_type in ("image", "pdf"):
            try:
                from app.services.design_tokens_applier_service import (
                    seed_design_tokens_gap_if_needed,
                )
                await seed_design_tokens_gap_if_needed(
                    self.db, project_id,
                    triggered_by_document_id=doc_id,
                )
                await self.db.commit()
            except Exception as exc:  # noqa: BLE001
                logger.warning(
                    "ingestion.design_tokens_seed_failed",
                    document_id=str(doc_id),
                    error=str(exc)[:300],
                )

        # MVP 24 Fase 24.2 — detector de PDF de questionário GCA.
        # Se o PDF é resposta do questionário técnico retroativo, aplicamos
        # as respostas **síncronamente** no GatekeeperItem em vez de mandar
        # pro pipeline Arguidor/LLM — é um PDF nosso, tem parsing canônico
        # determinístico, não precisa de LLM.
        questionnaire_applied: Optional[dict] = None
        if file_type == "pdf":
            try:
                from app.services.arguider_questionnaire_parser import (
                    apply_parsed_responses,
                    is_gca_questionnaire_pdf,
                    parse_questionnaire_pdf,
                )
                if is_gca_questionnaire_pdf(file_bytes):
                    parsed = parse_questionnaire_pdf(file_bytes)
                    report = await apply_parsed_responses(
                        self.db, project_id, uploaded_by, parsed,
                    )
                    # Marca o doc como questionário respondido — pipeline
                    # Arguidor/LLM não precisa reprocessar.
                    document.document_category = "arguider_questionnaire_response"
                    document.arguider_status = "completed"
                    document.arguider_completed_at = datetime.now(timezone.utc)
                    await self.db.commit()
                    questionnaire_applied = report.to_dict()
                    logger.info(
                        "ingestion.arguider_questionnaire_applied",
                        document_id=str(doc_id),
                        project_id=str(project_id),
                        **questionnaire_applied,
                    )
            except Exception as exc:
                # Não bloqueia upload se o detector/aplicador falhar —
                # degrada pro pipeline Celery padrão (fluxo LLM).
                logger.warning(
                    "ingestion.arguider_questionnaire_hook_failed",
                    document_id=str(doc_id),
                    error=str(exc)[:300],
                )

        if questionnaire_applied is not None:
            return {
                "document_id": str(doc_id),
                "status": "completed",
                "questionnaire_applied": questionnaire_applied,
                "message": (
                    f"Questionário técnico aplicado: "
                    f"{questionnaire_applied['applied']} itens resolvidos, "
                    f"{questionnaire_applied['skipped_blank']} em branco."
                ),
            }

        # Processamento sequencial canônico: doc nasce 'pending', dispatcher
        # com advisory lock garante 1-por-vez por projeto (sem race no upload
        # paralelo). Próximo doc é puxado pelo callback /ingestion-complete.
        # Questionnaire é sintético (Arq-M1) — não entra no pipeline.
        if file_type != "questionnaire":
            try:
                await dispatch_first_pending_for_project(self.db, project_id)
            except Exception as exc:  # noqa: BLE001
                logger.error(
                    "ingestion.dispatch_first_pending_failed",
                    document_id=str(doc_id),
                    project_id=str(project_id),
                    error=str(exc),
                    exc_info=True,
                )
                # Continua — watchdog (DT-073) reconsilia se ficou órfão

        logger.info(
            "ingestion.document_uploaded",
            document_id=str(doc_id),
            project_id=str(project_id),
            filename=original_filename,
            file_type=file_type,
            size=len(file_bytes),
        )

        return {
            "document_id": str(doc_id),
            "status": "pending",
            "message": "Documento recebido. Análise iniciada.",
        }

    @staticmethod
    def _valid_cpf(digits: str) -> bool:
        if len(digits) != 11 or digits == digits[0] * 11:
            return False
        s1 = sum(int(digits[i]) * (10 - i) for i in range(9))
        r1 = (s1 * 10) % 11
        d1 = 0 if r1 == 10 else r1
        if d1 != int(digits[9]):
            return False
        s2 = sum(int(digits[i]) * (11 - i) for i in range(10))
        r2 = (s2 * 10) % 11
        d2 = 0 if r2 == 10 else r2
        return d2 == int(digits[10])

    @staticmethod
    def _valid_cnpj(digits: str) -> bool:
        if len(digits) != 14 or digits == digits[0] * 14:
            return False
        w1 = [5, 4, 3, 2, 9, 8, 7, 6, 5, 4, 3, 2]
        w2 = [6, 5, 4, 3, 2, 9, 8, 7, 6, 5, 4, 3, 2]
        s1 = sum(int(digits[i]) * w1[i] for i in range(12))
        r1 = s1 % 11
        d1 = 0 if r1 < 2 else 11 - r1
        if d1 != int(digits[12]):
            return False
        s2 = sum(int(digits[i]) * w2[i] for i in range(13))
        r2 = s2 % 11
        d2 = 0 if r2 < 2 else 11 - r2
        return d2 == int(digits[13])

    @staticmethod
    def _valid_luhn(digits: str) -> bool:
        if len(digits) < 13 or len(digits) > 19:
            return False
        if digits == digits[0] * len(digits):
            return False  # todos-iguais (ex: "0000...") — Luhn aceita mas não é cartão real
        total = 0
        for i, ch in enumerate(reversed(digits)):
            n = int(ch)
            if i % 2 == 1:
                n *= 2
                if n > 9:
                    n -= 9
            total += n
        return total % 10 == 0

    @staticmethod
    def _extract_text_for_pii_scan(file_bytes: bytes, file_type: str) -> str:
        """DT-4 dogfood — Extrai texto humano-legível antes da triagem de PII.

        O scan antigo fazia `file_bytes.decode("utf-8", errors="ignore")`
        direto nos bytes — em PDFs/DOCX ricos em runs numéricos (xref
        tables, IDs internos, streams), isso ainda gerava janela de falso
        positivo mesmo após o reforço de DT-028 (mod-11/Luhn/contexto
        telefone). A correção estrutural é usar o extrator certo do
        formato antes de aplicar regex.

        - PDF: `extract_pdf_layered(bytes).text` (camadas AcroForm + texto).
        - DOCX: `extract_rich_text(bytes)` (parágrafos + tabelas).
        - Outros (texto/markdown/json/csv etc): decode UTF-8 direto.
        - image/spreadsheet: retorna "" (não escaneia).

        Trunca em 100 KB pra não pagar pelo tempo de regex em docs gigantes.
        """
        if file_type in ("image", "spreadsheet"):
            return ""

        try:
            if file_type == "pdf":
                from app.services.pdf_layered_extractor import extract_pdf_layered
                text = extract_pdf_layered(file_bytes).text or ""
            elif file_type == "docx":
                from app.services.rich_docx_extractor import extract_rich_text
                text = extract_rich_text(file_bytes) or ""
                # extract_rich_text começa com `[Erro ...]` em falha — não escaneia
                if text.startswith("["):
                    text = ""
            else:
                text = file_bytes.decode("utf-8", errors="ignore")
        except Exception:
            # Qualquer falha do extractor → cai no caminho seguro (string vazia,
            # zero PII detectado). PII detection nunca deve quebrar a ingestão.
            return ""

        return text[:100_000]

    @staticmethod
    def _detect_pii(file_bytes: bytes, file_type: str) -> tuple[bool, list[str]]:
        """Triagem básica de PII em conteúdo textual.

        Para CPF/CNPJ/cartão: exige que o valor passe pelo dígito verificador real
        (mod-11 / Luhn) — regex puro dá falso-positivo em PDFs/binários que têm
        runs de 14 dígitos em xref tables, IDs de objetos etc.

        DT-4 dogfood: extração de texto agora é via extractor real do formato
        (`_extract_text_for_pii_scan`), não mais decode("utf-8") cru sobre
        bytes. Elimina falso-positivo em PDFs/DOCX ricos em runs numéricos.
        """
        import re

        text = IngestionService._extract_text_for_pii_scan(file_bytes, file_type)
        if not text:
            return False, []

        # Remove UUIDs antes do scan: fragments de hex/digit dentro de
        # UUIDs podem casar com regex de cartão/CPF/CNPJ por coincidência
        # (e Luhn/mod-11 aceitam ~10% dos números aleatórios). Caso real
        # 2026-05-04 BRT: pfq-id `...41ad-9042-186801726975` foi parsed
        # como cartão de crédito válido, mandando doc legítimo pra
        # quarentena. Substituir por espaço preserva offsets sem afetar
        # outros detectores que usam contexto de palavra.
        text = re.sub(
            r"\b[0-9a-f]{8}-[0-9a-f]{4}-[0-9a-f]{4}-[0-9a-f]{4}-[0-9a-f]{12}\b",
            " ",
            text,
            flags=re.IGNORECASE,
        )

        only_digits = re.compile(r"\D")
        detected: list[str] = []

        # CPF: 11 dígitos, opcional formatação, valida mod-11
        for m in re.finditer(r"\b\d{3}\.?\d{3}\.?\d{3}-?\d{2}\b", text):
            if IngestionService._valid_cpf(only_digits.sub("", m.group(0))):
                detected.append("cpf")
                break

        # CNPJ: 14 dígitos, opcional formatação, valida mod-11
        for m in re.finditer(r"\b\d{2}\.?\d{3}\.?\d{3}/?\d{4}-?\d{2}\b", text):
            if IngestionService._valid_cnpj(only_digits.sub("", m.group(0))):
                detected.append("cnpj")
                break

        # Cartão de crédito: 13-19 dígitos, valida Luhn
        for m in re.finditer(r"\b\d{4}[\s-]?\d{4}[\s-]?\d{4}[\s-]?\d{4}\b", text):
            if IngestionService._valid_luhn(only_digits.sub("", m.group(0))):
                detected.append("cartao_credito")
                break

        # Email pessoal: regex é suficiente (formato é discriminante)
        if re.search(r"\b[a-zA-Z0-9._%+-]+@(gmail|hotmail|yahoo|outlook)\.[a-zA-Z]{2,}\b", text):
            detected.append("email_pessoal")

        # Telefone BR: regex antigo (`\b\(?\d{2}\)?\s?\d{4,5}-?\d{4}\b`)
        # era promíscuo — matchava qualquer sequência 2+4-5+4 dígitos sem
        # contexto, dando falso-positivo em PDFs com métricas, IDs,
        # coordenadas, timestamps. DT-028 quitada: agora exige padrão
        # formatado inequívoco OU contexto explícito ("tel"/"fone"/
        # "whatsapp" nas proximidades), e valida DDD como 11-99.
        if IngestionService._has_br_phone(text):
            detected.append("telefone_br")

        return len(detected) > 0, detected

    @staticmethod
    def _has_br_phone(text: str) -> bool:
        """Detecção estrita de telefone BR.

        Aceita apenas padrões inequívocos:
          (a) "(DD) 9NNNN-NNNN" ou "(DD) NNNN-NNNN"  — parênteses no DDD
          (b) "+55 DD 9NNNN-NNNN"                    — E.164
          (c) "DD 9NNNN-NNNN" / "DD NNNN-NNNN"       — hífen obrigatório
          (d) qualquer match acima com contexto "tel"/"fone"/"telefone"/
              "whatsapp"/"celular"/"cel." numa janela de 40 chars antes.

        Rejeita:
          - sequências sem hífen nem parênteses (lookout de PDF binário);
          - DDs fora de 11-99 (faixa válida dos DDDs no Brasil);
          - números com todos dígitos iguais (1111111111) ou sentinelas.
        """
        import re as _re

        # Padrões inequívocos — exige pelo menos 1 separador (parêntese
        # ou hífen) no formato esperado do Brasil.
        patterns = [
            # (DD) NNNNN-NNNN ou (DD) NNNN-NNNN — parênteses obrigatórios
            r"\((\d{2})\)\s?9?\d{4}-\d{4}\b",
            # +55 DD NNNNN-NNNN — E.164
            r"\+55\s?(\d{2})\s?9?\d{4}-?\d{4}\b",
            # DD NNNNN-NNNN — hífen final obrigatório (elimina runs crus)
            r"\b(\d{2})\s\d{4,5}-\d{4}\b",
        ]

        def _valid_ddd(ddd_str: str) -> bool:
            try:
                ddd = int(ddd_str)
                return 11 <= ddd <= 99
            except ValueError:
                return False

        def _not_sentinel(raw: str) -> bool:
            digits = _re.sub(r"\D", "", raw)
            # Rejeita só "todos iguais" (1111111111, 0000000000).
            # Ter 2 dígitos distintos (ex.: "1199999999") é normal em
            # telefone real.
            if len(set(digits)) < 2:
                return False
            return True

        for pat in patterns:
            for m in _re.finditer(pat, text):
                if _valid_ddd(m.group(1)) and _not_sentinel(m.group(0)):
                    return True

        # Fallback com contexto: se o texto tem "tel"/"fone"/"celular"/
        # "whatsapp" próximo a uma sequência plausível, também conta.
        ctx_re = _re.compile(
            r"(?:tel|telefone|fone|cel(?:ular)?|whats?app|contato)"
            r"[^\n]{0,40}?(\d{2}[\s().-]{0,3}9?\d{4}[\s.-]?\d{4})",
            _re.IGNORECASE,
        )
        for m in ctx_re.finditer(text):
            digits = _re.sub(r"\D", "", m.group(1))
            if len(digits) in (10, 11) and _not_sentinel(digits):
                return True

        return False

    # MVP 8 Fase 1 — Mapa canônico de estágio → porcentagem.
    # Ordem: queued(0) → extracting_text(10) → analyzing(40) →
    # updating_ocg(70) → regenerating_backlog(90) → completed(100).
    _STAGE_PERCENTS: dict[str, int] = {
        "queued": 0,
        "extracting_text": 10,
        "analyzing": 40,
        "updating_ocg": 70,
        "regenerating_backlog": 90,
        "completed": 100,
        "failed": 0,  # percent não regride; frontend mostra último valor
    }

    @classmethod
    async def _update_stage(
        cls,
        db,
        document_id: UUID,
        stage: str,
        *,
        percent: int | None = None,
    ) -> None:
        """MVP 8 Fase 1 — atualiza arguider_stage + porcentagem do documento.

        Commit dedicado para que o frontend (polling) enxergue o estágio
        mesmo que o resto do pipeline demore. Se `percent` não vem, usa
        o default do mapa. Nunca regride porcentagem exceto quando
        explicitamente passado (ex: reset pra 0 em reanalyze).
        """
        from datetime import datetime as _dt, timezone as _tz
        from app.models.base import IngestedDocument as _Doc
        doc = await db.get(_Doc, document_id)
        if not doc:
            logger.warning(
                "ingestion.update_stage_doc_not_found",
                document_id=str(document_id),
                stage=stage,
            )
            return
        doc.arguider_stage = stage
        if doc.arguider_status == "pending":
            doc.arguider_status = "processing"
        if percent is None:
            percent = cls._STAGE_PERCENTS.get(stage, doc.arguider_progress_percent)
        # Não regride (exceto se caller explicitamente passar valor menor)
        if percent >= doc.arguider_progress_percent or stage == "failed":
            doc.arguider_progress_percent = percent
        doc.arguider_stage_updated_at = _dt.now(_tz.utc)
        try:
            await db.commit()
        except Exception as _commit_exc:
            # DT-072: session pode estar em estado ambíguo depois de
            # fluxos longos (ex: OCGUpdaterService faz múltiplos commits
            # na mesma session). Loga e propaga pro caller decidir se
            # tenta de novo com session fresca.
            logger.warning(
                "ingestion.update_stage_commit_failed",
                document_id=str(document_id),
                stage=stage,
                error=str(_commit_exc),
            )
            raise

    @staticmethod
    def _deliverable_category_from_module(module_type: str | None) -> str:
        """Mapeia categoria canônica do MVP 9 pra schema de DELIVERABLES."""
        m = (module_type or "").lower()
        return {
            "infrastructure": "config",
            "deploy_pipeline": "config",
            "observability": "config",
            "middleware": "code",
            "backend_service": "code",
            "feature": "code",
        }.get(m, "other")

    @classmethod
    async def _link_target_module_after_pipeline(
        cls,
        *,
        document_id: UUID,
        project_id: UUID,
    ) -> None:
        """MVP 9 Fase 9.5.2 — após o pipeline completar com sucesso, se o doc
        está vinculado a um item do Roadmap (target_module_id), transita o
        item pra `adicionado` E cria row em `project_deliverables`.

        MVP 9 Fase 9.3 (extension): após transição, dispara orquestração
        Premium pra avaliar readiness_status. Falha silenciosa se Premium
        não configurado (não invalida fluxo principal).

        Wrapper com session fresca (isola do estado da session do
        _analyze_async). A lógica core está em
        `_link_target_module_in_session` pra ser testável.
        """
        from app.db.database import AsyncSessionLocal as _ASL
        async with _ASL() as _db:
            target_module_id = await cls._link_target_module_in_session(
                _db, document_id, project_id,
            )
            await _db.commit()

        # MVP 9 Fase 9.3 — avalia readiness em sessão própria pra
        # evitar dependência da session que acabou de commitar.
        if target_module_id:
            await cls._evaluate_readiness_safe(project_id, target_module_id)

    @classmethod
    async def _evaluate_readiness_safe(
        cls, project_id: UUID, module_id: UUID,
    ) -> None:
        """Wrapper que invoca orquestração Premium e captura falhas
        operacionais (sem Premium configurado, timeout, etc) sem
        derrubar o fluxo. Erros viram log warning."""
        from app.db.database import AsyncSessionLocal as _ASL
        from app.services.module_orchestration_service import evaluate_module_readiness
        try:
            async with _ASL() as _db:
                await evaluate_module_readiness(_db, project_id, module_id)
                logger.info(
                    "ingestion.readiness_evaluated",
                    module_id=str(module_id), project_id=str(project_id),
                )
        except RuntimeError as exc:
            # Sem Premium configurado — esperado em projetos só com Ollama
            logger.info(
                "ingestion.readiness_skipped_no_premium",
                module_id=str(module_id), reason=str(exc),
            )
        except Exception as exc:
            logger.warning(
                "ingestion.readiness_eval_failed",
                module_id=str(module_id), error=str(exc),
            )

    @classmethod
    async def _link_target_module_in_session(
        cls,
        db,
        document_id: UUID,
        project_id: UUID,
    ) -> UUID | None:
        """Lógica core da Fase 9.5.2 testável com session arbitrária.

        Retorna o `module_id` vinculado quando aplicou transição/criou
        deliverable, ou `None` quando doc não tinha vínculo. Caller
        usa o retorno pra disparar hooks subsequentes (Fase 9.3
        readiness eval).

        Idempotente: chamadas duplicadas no mesmo (doc, módulo) não
        duplicam deliverable nem mudam status repetido. Não comita —
        caller decide.
        """
        from app.models.base import (
            IngestedDocument as _Doc,
            ModuleCandidate as _MC,
            ProjectDeliverable as _Del,
        )
        from app.constants.module_categories import (
            CATEGORY_LABELS_PT_BR, is_allowed_transition,
        )
        import re as _re

        doc = await db.get(_Doc, document_id)
        if not doc or not doc.target_module_id:
            return None  # Doc sem vínculo — nada a fazer

        mc = await db.get(_MC, doc.target_module_id)
        if not mc or mc.project_id != project_id:
            logger.warning(
                "ingestion.target_module_not_found_or_cross_project",
                document_id=str(document_id),
                target_module_id=str(doc.target_module_id),
            )
            return None

        # Transição: status atual → 'adicionado' (se permitido)
        target_status = "adicionado"
        current = mc.status or "sugerido"
        if current != target_status and is_allowed_transition(current, target_status):
            mc.status = target_status
            logger.info(
                "ingestion.module_status_transitioned",
                module_id=str(mc.id), project_id=str(project_id),
                from_status=current, to_status=target_status,
            )
        elif current != target_status:
            logger.warning(
                "ingestion.module_status_transition_blocked",
                module_id=str(mc.id), from_status=current,
                to_status=target_status,
            )

        # Cria DELIVERABLE idempotente (UniqueConstraint por
        # project_id + normalized_name).
        normalized = _re.sub(r"\s+", " ", (mc.name or "").strip()).lower()[:500]
        existing = await db.execute(
            select(_Del).where(
                _Del.project_id == project_id,
                _Del.normalized_name == normalized,
            )
        )
        existing_row = existing.scalar_one_or_none()
        cat_label = CATEGORY_LABELS_PT_BR.get(mc.module_type or "", "Outro")
        if existing_row is None:
            db.add(_Del(
                project_id=project_id,
                name=mc.name or "Item sem nome",
                normalized_name=normalized,
                category=IngestionService._deliverable_category_from_module(mc.module_type),
                kind=f"roadmap_module:{mc.module_type or 'feature'}",
                status="declared",
                notes=f"Criado automaticamente ao adicionar item do Roadmap ({cat_label}). Doc fonte: {doc.original_filename}",
            ))
            logger.info(
                "ingestion.deliverable_created_from_module",
                module_id=str(mc.id), project_id=str(project_id),
                deliverable_name=mc.name,
            )
        else:
            logger.info(
                "ingestion.deliverable_already_exists",
                module_id=str(mc.id), normalized_name=normalized,
            )

        return mc.id

    @classmethod
    async def _update_stage_fresh(
        cls,
        document_id: UUID,
        stage: str,
        *,
        percent: int | None = None,
    ) -> None:
        """DT-072 — versão de `_update_stage` que SEMPRE usa session
        nova. Usar nos pontos finais do pipeline, onde a session
        corrente pode ter sido corrompida por serviços que fazem
        commits internos (OCGUpdaterService). Evita o estado "stage
        parado em updating_ocg/70%" visto no dogfood 2026-04-19.
        """
        from app.db.database import AsyncSessionLocal as _ASL
        try:
            async with _ASL() as _fresh:
                await cls._update_stage(_fresh, document_id, stage, percent=percent)
        except Exception as exc:
            logger.warning(
                "ingestion.update_stage_fresh_failed",
                document_id=str(document_id),
                stage=stage,
                error=str(exc),
            )

    async def _notify_provider_fallback(
        self,
        db,
        *,
        project_id: UUID,
        primary_provider: str,
        fallback_provider: str,
        document_id: UUID,
    ) -> None:
        """DT-064 — Avisa GPs + Admins do projeto que o provider IA
        primário falhou e o sistema fez fallback automático para o
        próximo da cadeia. Inclui aviso sobre possível aumento de tempo
        de processamento quando o fallback é provider local (Ollama)."""
        from sqlalchemy import select as _select, or_ as _or
        from app.models.base import ProjectMember, User
        from app.services.notification_inapp_service import InAppNotificationService

        # Labels amigáveis + aviso sobre tempo
        _names = {
            "anthropic": "Anthropic (Claude)",
            "openai": "OpenAI (GPT)",
            "gemini": "Google (Gemini)",
            "deepseek": "DeepSeek",
            "grok": "Grok",
            "ollama": "Ollama (IA local)",
        }
        primary_name = _names.get(primary_provider, primary_provider)
        fallback_name = _names.get(fallback_provider, fallback_provider)

        is_local_fallback = (fallback_provider == "ollama")
        if is_local_fallback:
            time_hint = (
                " O processamento pode ficar mais lento que o habitual, pois a "
                "IA local tem latência maior que os provedores remotos."
            )
        else:
            time_hint = (
                " O tempo de processamento pode variar conforme o provedor de "
                "fallback."
            )

        title = f"Fallback de IA ativado: usando {fallback_name}"
        message = (
            f"O provedor principal ({primary_name}) não respondeu com sucesso "
            f"na análise do documento. O sistema trocou automaticamente para "
            f"{fallback_name} para não travar o fluxo.{time_hint} "
            f"Se o problema persistir, verifique credenciais e cotas em "
            f"Configurações do projeto → IA."
        )

        # Destinatários: GPs do projeto + Admins ativos
        members_q = _select(ProjectMember.user_id).where(
            ProjectMember.project_id == project_id,
            ProjectMember.role == "gp",
            ProjectMember.accepted_at.isnot(None),
            ProjectMember.is_active.is_(True),
        )
        admins_q = _select(User.id).where(
            User.is_admin.is_(True), User.is_active.is_(True),
        )
        gp_ids = list((await db.execute(members_q)).scalars().all())
        admin_ids = list((await db.execute(admins_q)).scalars().all())
        recipients = list({*gp_ids, *admin_ids})

        notif = InAppNotificationService(db)
        for uid in recipients:
            try:
                await notif.notify(
                    user_id=uid,
                    event_type="ai.provider_fallback",
                    title=title,
                    message=message,
                    project_id=project_id,
                    resource_type="ingested_document",
                    resource_id=document_id,
                    link=f"/projects/{project_id}/ingestion",
                    severity="warning",
                )
            except Exception as e:  # noqa: BLE001
                logger.warning("ingestion.fallback_notification_failed",
                               user_id=str(uid), error=str(e))

        logger.info(
            "ingestion.provider_fallback_notified",
            project_id=str(project_id),
            document_id=str(document_id),
            primary=primary_provider,
            fallback=fallback_provider,
            recipients=len(recipients),
        )

    # DT-3 dogfood: tempo máximo de uma análise antes de matar a task.
    # Análises reais com Anthropic levam segundos a minutos; com Ollama
    # local sobem pra alguns minutos. 10 min é margem confortável.
    _ANALYZE_TIMEOUT_SECONDS = 600

    async def _analyze_with_timeout(
        self,
        document_id: UUID,
        project_id: UUID,
        file_bytes: bytes,
        file_type: str,
    ):
        """DT-3 dogfood — Wrapper que aplica timeout duro a `_analyze_async`.

        Em caso de TimeoutError, marca o doc como `arguider_status='error'`
        com mensagem clara. Sem isso, doc fica preso em 'processing' até
        watchdog do startup limpar (que só roda se o backend reiniciar).
        """
        import asyncio as _asyncio
        try:
            await _asyncio.wait_for(
                self._analyze_async(document_id, project_id, file_bytes, file_type),
                timeout=self._ANALYZE_TIMEOUT_SECONDS,
            )
        except _asyncio.TimeoutError:
            logger.error(
                "ingestion.analyze_timeout",
                document_id=str(document_id),
                project_id=str(project_id),
                timeout_s=self._ANALYZE_TIMEOUT_SECONDS,
            )
            try:
                from app.db.database import AsyncSessionLocal as _ASL
                from app.models.base import IngestedDocument as _Doc
                async with _ASL() as _db:
                    _doc = await _db.get(_Doc, document_id)
                    if _doc and _doc.arguider_status == "processing":
                        _doc.arguider_status = "error"
                        _doc.arguider_error_message = (
                            f"Análise excedeu o tempo máximo de "
                            f"{self._ANALYZE_TIMEOUT_SECONDS // 60} minutos. "
                            "Tente novamente ou reduza o tamanho do documento."
                        )
                        _doc.arguider_stage = "failed"
                        await _db.commit()
            except Exception as _e:
                logger.warning(
                    "ingestion.analyze_timeout_cleanup_failed",
                    document_id=str(document_id), error=str(_e),
                )

    async def _analyze_async(
        self,
        document_id: UUID,
        project_id: UUID,
        file_bytes: bytes,
        file_type: str,
    ):
        """Executa análise do Arguidor em background.

        DT-064 — Fallback automático entre providers IA: se o provider
        default falhar (rate limit, quota, 5xx, timeout, erro de rede
        em geral), tenta o próximo provider configurado no projeto em
        cascata. Notifica GPs + Admins quando o fallback é acionado,
        avisando sobre possível aumento de tempo se cair em IA local.
        """
        try:
            from app.db.database import AsyncSessionLocal
            async with AsyncSessionLocal() as db:
                # DT-064: resolver toda a cadeia de providers configurados
                # (default primeiro + demais validados). Loop vai tentar em
                # sequência se o primeiro falhar com erro transiente.
                from app.services.ai_key_resolver import AIKeyResolver
                provider_chain = await AIKeyResolver.resolve_project_provider_chain(db, project_id)

                if not provider_chain:
                    raise RuntimeError(
                        "Nenhum provedor de IA configurado no projeto. "
                        "GP deve configurar provedor e chave em Settings > IA."
                    )

                # MVP 8 Fase 3B — DocumentExtractor ganha contexto do
                # projeto pra ativar OCR via Vision quando PDFs escaneados
                # não produzem texto nas camadas 1+2.
                extractor = DocumentExtractor(project_id=project_id, db=db)

                # MVP 8 Fase 1 — marcar estágio "extracting_text"
                await IngestionService._update_stage(db, document_id, "extracting_text")

                # Extrair texto (fora do loop — não depende do provider)
                doc_text = await extractor.extract_text(file_bytes, file_type)

                # MVP 29 Fase 3 — gerar DocumentCanonical pra enviar ao
                # Arguidor dirigido (3-5× menos tokens que texto bruto).
                # Best-effort: se falhar, cai no texto bruto original.
                doc_canonical = None
                canonical_doc_type = {
                    "md": "MD", "pdf": "PDF", "docx": "DOCX",
                }.get((file_type or "").lower())
                if canonical_doc_type:
                    try:
                        from app.services.document_canonicalizer import canonicalize
                        _doc_row = await db.get(IngestedDocument, document_id)
                        _orig_name = _doc_row.original_filename if _doc_row else ""
                        doc_canonical = canonicalize(
                            file_bytes, _orig_name, canonical_doc_type,
                        )
                        logger.info(
                            "ingestion.canonicalized",
                            document_id=str(document_id),
                            document_type=canonical_doc_type,
                            sections=len(doc_canonical.sections),
                            requirements=len(doc_canonical.requirements),
                            raw_chars=len(doc_text or ""),
                        )
                    except Exception as exc:  # noqa: BLE001
                        logger.warning(
                            "ingestion.canonicalize_failed",
                            document_id=str(document_id),
                            error=str(exc),
                            fallback="raw_text",
                        )
                        doc_canonical = None

                # MVP 8 Fase 1 — texto extraído, próximo estágio será "analyzing"
                # (a atualização real acontece dentro do loop, antes do
                # arguider.analyze_document)

                # Buscar OCG + análises prévias (fora do loop também)
                ocg_result = await db.execute(
                    select(OCG).where(OCG.project_id == project_id).order_by(OCG.created_at.desc()).limit(1)
                )
                _ocg = ocg_result.scalar_one_or_none()
                _current_ocg = json.loads(_ocg.ocg_data) if _ocg and _ocg.ocg_data else {}
                _prev_analyses_raw = await db.execute(
                    select(ArguiderAnalysis).where(ArguiderAnalysis.project_id == project_id)
                )
                _prev_analyses = []
                for a in _prev_analyses_raw.scalars().all():
                    try:
                        _prev_analyses.append({
                            "document_classification": json.loads(a.document_classification),
                            "gaps": json.loads(a.gaps),
                            "module_candidates": json.loads(a.module_candidates),
                        })
                    except json.JSONDecodeError:
                        pass

                # ═══ DT-064: Loop de fallback entre providers ═══
                last_error = None
                successful_provider = None
                for idx, pcfg in enumerate(provider_chain):
                    provider = pcfg["provider"]
                    model = pcfg.get("model")
                    base_url = pcfg.get("base_url") if provider == "ollama" else None

                    try:
                        project_api_key = await AIKeyResolver.get_project_key(
                            db, project_id, provider=provider,
                        )
                    except Exception as e:
                        logger.warning("ingestion.key_resolve_failed",
                                       provider=provider, error=str(e))
                        last_error = f"Chave do provider {provider} indisponível: {e}"
                        continue

                    logger.info("ingestion.provider_attempt",
                                document_id=str(document_id),
                                attempt=idx + 1,
                                total=len(provider_chain),
                                provider=provider,
                                model=model or "(default)")

                    try:
                        # FASE 1 Refactor — AuditorOrchestratorService
                        # Replace Arguidor with new orchestrator pipeline:
                        # 1. Chunk documento → DocumentRouteMap
                        # 2. Auditor análise inicial → AuditorOutput
                        # 3. 7 personas em paralelo → personas_responses
                        # 4. OCGConsolidator → OCG updates + conflicts
                        from app.services.auditor_orchestrator_service import AuditorOrchestratorService
                        from app.services.llm_client import create_llm_client

                        await IngestionService._update_stage(db, document_id, "analyzing")

                        llm_client = create_llm_client(
                            provider=provider,
                            api_key=project_api_key,
                            model=model,
                            base_url=base_url,
                        )

                        orchestrator = AuditorOrchestratorService(db, llm_client)
                        orchestration_result = await orchestrator.orchestrate(
                            document_id=document_id,
                            project_id=project_id,
                            document_text=doc_text,
                            file_type=file_type,
                        )

                        if not orchestration_result.get("success"):
                            raise RuntimeError(
                                f"Orchestration failed: {orchestration_result.get('error')}"
                            )

                        successful_provider = provider
                        if idx > 0:
                            logger.warning(
                                "ingestion.provider_fallback_succeeded",
                                document_id=str(document_id),
                                primary_provider=provider_chain[0]["provider"],
                                fallback_provider=provider,
                                providers_tried=idx + 1,
                            )
                            # DT-064: notificar GPs + admins do projeto que o
                            # fallback aconteceu e que o processamento pode
                            # estar mais lento se o fallback for provider local.
                            await self._notify_provider_fallback(
                                db,
                                project_id=project_id,
                                primary_provider=provider_chain[0]["provider"],
                                fallback_provider=provider,
                                document_id=document_id,
                            )
                        break  # sucesso — sai do loop
                    except Exception as e:
                        err_msg = str(e)
                        last_error = err_msg
                        should_fallback = AIKeyResolver.should_fallback_to_next_provider(err_msg)
                        logger.warning(
                            "ingestion.provider_failed",
                            document_id=str(document_id),
                            provider=provider,
                            should_fallback=should_fallback,
                            remaining_providers=len(provider_chain) - idx - 1,
                            error=err_msg[:200],
                        )
                        if not should_fallback:
                            # Erro de parâmetro/schema — fallback não resolve.
                            # Aborta cascata e propaga.
                            raise
                        # Reset do status do documento para próxima tentativa.
                        # MVP 29 Fase 29.1: zera `arguider_started_at` junto
                        # pra não deixar o doc em estado zombie (pending com
                        # started_at preenchido) se o worker morrer entre o
                        # reset e o retry. O próximo provider/retry vai setar
                        # started_at novamente via `_update_stage` canônico.
                        from app.models.base import IngestedDocument as _IngDoc
                        _doc = await db.get(_IngDoc, document_id)
                        if _doc:
                            _doc.arguider_status = "pending"
                            _doc.arguider_started_at = None
                            _doc.arguider_error_message = None
                            await db.commit()
                        continue  # tenta próximo provider

                if not successful_provider:
                    # Todos os providers falharam com erro transiente.
                    # Propaga para o caller ser marcado como erro.
                    raise RuntimeError(
                        f"Todos os {len(provider_chain)} providers configurados "
                        f"retornaram erro transiente. Último erro: {last_error[:300] if last_error else 'desconhecido'}"
                    )

                # Marcar documento como analisado com OCG atualizado
                doc = await db.get(IngestedDocument, document_id)
                if doc and doc.arguider_status == "completed":
                    doc.ocg_updated = True
                    await db.commit()

                    # Registrar evento DOCUMENT_INGESTED
                    from app.services.audit_service import AuditService
                    audit = AuditService(db)
                    await audit.log_event(
                        event_type="DOCUMENT_INGESTED",
                        resource_type="ingested_document",
                        resource_id=document_id,
                        details={
                            "project_id": str(project_id),
                            "filename": doc.original_filename,
                            "file_type": doc.file_type,
                            "ocg_updated": True,
                        },
                    )
                    await db.commit()

                    logger.info("ingestion.analysis_complete_ocg_updated",
                               document_id=str(document_id),
                               project_id=str(project_id))

                    # === OCG REATIVO: Atualizar OCG via IA ===
                    try:
                        from app.services.ocg_updater_service import OCGUpdaterService
                        import json as _json

                        # Carregar análise do Arguidor
                        arguider_result = await db.execute(
                            select(ArguiderAnalysis).where(
                                (ArguiderAnalysis.document_id == document_id) &
                                (ArguiderAnalysis.project_id == project_id)
                            )
                        )
                        analysis = arguider_result.scalar_one_or_none()
                        analysis_data = {}
                        if analysis:
                            try:
                                analysis_data = {
                                    "classification": _json.loads(analysis.document_classification) if analysis.document_classification else {},
                                    "gaps": _json.loads(analysis.gaps) if analysis.gaps else [],
                                    "module_candidates": _json.loads(analysis.module_candidates) if analysis.module_candidates else [],
                                }
                            except _json.JSONDecodeError:
                                pass

                        # MVP 8 Fase 1 — estágio "updating_ocg"
                        await IngestionService._update_stage(db, document_id, "updating_ocg")

                        # DT-065 — OCG updater também participa da cadeia
                        # de fallback. Se o provider atual falhar no
                        # update do OCG, tentar os próximos providers
                        # SEM refazer a análise do Arguidor (preservada
                        # em arguider_analyses).
                        update_result = None
                        ocg_last_error = None
                        ocg_successful_provider = None
                        for ocg_idx, ocg_pcfg in enumerate(provider_chain):
                            ocg_provider = ocg_pcfg["provider"]
                            # OCG Updater hoje lê provider via
                            # AIKeyResolver._resolve_project_provider (que
                            # só retorna o default). Para forçar o provider
                            # corrente no OCG updater, setamos
                            # temporariamente o default na sessão. Ao fim
                            # do loop (ou no sucesso) restauramos.
                            try:
                                updater = OCGUpdaterService(db)
                                # Se o provider do OCG diverge do provider
                                # que acabou de ser usado pelo Arguidor,
                                # vamos precisar reinjetar na sessão uma
                                # override. Mas o _call_llm do updater já
                                # resolve do DB — solução simples: alterar
                                # temporariamente o settings_json ficaria
                                # invasivo. Em vez disso, se idx!=ocg_idx,
                                # usamos try/except e caímos no fallback.
                                update_result = await updater.update_ocg_from_arguider(
                                    project_id=project_id,
                                    arguider_analysis=analysis_data,
                                    document_id=document_id,
                                    actor_id=doc.uploaded_by if doc else None,
                                    trigger_source="document_ingestion",
                                )

                                # DT-AUDITORIA-002: "awaiting_ocg" significa
                                # que OCG (legacy, do questionário) não está
                                # disponível. Isso é ok — em Phase B (personas
                                # diretas via documento), pode não haver OCG.
                                # Log e continua (não bloqueia completion).
                                if update_result and update_result.get("status") == "awaiting_ocg":
                                    logger.info(
                                        "ingestion.awaiting_ocg",
                                        document_id=str(document_id),
                                        project_id=str(project_id),
                                        retry_at=update_result.get("retry_at"),
                                    )

                                ocg_successful_provider = ocg_provider
                                if ocg_idx > 0:
                                    logger.warning(
                                        "ingestion.ocg_fallback_succeeded",
                                        document_id=str(document_id),
                                        fallback_provider=ocg_provider,
                                        attempts=ocg_idx + 1,
                                    )
                                break
                            except Exception as _e:
                                ocg_last_error = str(_e)
                                should_fb = AIKeyResolver.should_fallback_to_next_provider(ocg_last_error)
                                logger.warning(
                                    "ingestion.ocg_provider_failed",
                                    document_id=str(document_id),
                                    provider=ocg_provider,
                                    should_fallback=should_fb,
                                    error=ocg_last_error[:200],
                                )
                                if not should_fb or ocg_idx == len(provider_chain) - 1:
                                    # Sem fallback possível — propaga erro
                                    # para o except externo (log + stage
                                    # completed@95%).
                                    raise

                        # Propagar se houve mudanças — fire-and-forget em
                        # task separada com SUA PRÓPRIA sessão. Isola falhas
                        # de propagação do commit do OCG (que já aconteceu)
                        # e libera o caller para retornar imediatamente.
                        if update_result and update_result.get("changes"):
                            # MVP 8 Fase 1 — estágio "regenerating_backlog"
                            # DT-072: session fresca. A session `db` passou pelo
                            # OCGUpdaterService que faz múltiplos commits
                            # internos; depois dele, `db.commit()` direto
                            # silenciou sem persistir stage.
                            await IngestionService._update_stage_fresh(document_id, "regenerating_backlog")
                            # MVP 13 Fase 13.3b: Celery tasks com retry + ACK late.
                            from app.tasks.pipeline import (
                                propagate_task,
                                reevaluate_gatekeeper_task,
                            )
                            propagate_task.delay(
                                str(project_id),
                                list(update_result["changes"]),
                                update_result.get("version_to"),
                            )
                            # Reavaliar Gatekeeper com o OCG novo (MVP 2 §10).
                            # Também fire-and-forget via Celery: grava evento
                            # GATEKEEPER_REEVALUATED no audit_log com
                            # blocking_pillars/derived_status resultantes.
                            reevaluate_gatekeeper_task.delay(
                                str(project_id),
                                update_result.get("version_to"),
                                "document_ingestion",
                            )

                        logger.info("ingestion.ocg_reactive_complete",
                                   document_id=str(document_id),
                                   ocg_updated=bool(update_result))

                        # MVP 8 Fase 1 — marcar estágio "completed" após
                        # sucesso completo do pipeline (arguider + ocg + propagação disparada).
                        # DT-072: session fresca (mesmo motivo de _update_stage_fresh acima).
                        await IngestionService._update_stage_fresh(document_id, "completed", percent=100)

                        # MVP 9 Fase 9.5.2 — se este doc estava vinculado a
                        # um item do Roadmap (target_module_id), transita o
                        # item pra `adicionado` e cria DELIVERABLE automático.
                        # Isolado em try/except: falha aqui não invalida o
                        # pipeline completo — log warning + segue.
                        try:
                            await IngestionService._link_target_module_after_pipeline(
                                document_id=document_id,
                                project_id=project_id,
                            )
                        except Exception as _link_exc:
                            logger.warning(
                                "ingestion.target_module_link_failed",
                                document_id=str(document_id),
                                error=str(_link_exc),
                            )

                    except Exception as e:
                        import traceback
                        logger.warning(
                            "ingestion.ocg_reactive_error",
                            document_id=str(document_id),
                            error=str(e) or repr(e),
                            error_type=type(e).__name__,
                            traceback=traceback.format_exc(),
                        )
                        # Mesmo com falha no OCG reativo, o Arguidor já
                        # completou — marcamos completed mas com porcentagem
                        # ligeiramente menor pro frontend mostrar o alerta
                        # via arguider_error_message (que vem do caller).
                        # DT-072: session fresca (a original pode ter
                        # rollbackado com a exception).
                        await IngestionService._update_stage_fresh(document_id, "completed", percent=95)

        except Exception as e:
            logger.error("ingestion.analysis_async_error", document_id=str(document_id), error=str(e))
            # MVP 8 Fase 1 — marcar estágio "failed" para frontend parar o polling.
            # Fix 2026-04-25: também marcar arguider_status='error' + mensagem,
            # senão o doc fica preso em 'processing' eternamente (sem timeout).
            try:
                from app.db.database import AsyncSessionLocal as _ASL
                from app.models.base import IngestedDocument as _Doc
                from datetime import datetime as _dt, timezone as _tz
                async with _ASL() as _db:
                    _doc = await _db.get(_Doc, document_id)
                    if _doc:
                        _doc.arguider_status = "error"
                        _doc.arguider_stage = "failed"
                        _doc.arguider_completed_at = _dt.now(_tz.utc)
                        if not _doc.arguider_error_message:
                            _doc.arguider_error_message = (str(e) or repr(e))[:500]
                        await _db.commit()
            except Exception:
                pass

    async def list_documents(self, project_id: UUID) -> list[dict]:
        """Lista documentos do projeto com tokens_used da análise Arguidor."""
        from sqlalchemy import outerjoin

        # LEFT JOIN com ArguiderAnalysis pra pegar tokens_used.
        # MVP 34: filtra docs soft-deleted — UI lista apenas docs ativos.
        result = await self.db.execute(
            select(IngestedDocument, ArguiderAnalysis.tokens_used)
            .outerjoin(ArguiderAnalysis, IngestedDocument.id == ArguiderAnalysis.document_id)
            .where(IngestedDocument.project_id == project_id)
            .where(IngestedDocument.deleted_at.is_(None))
            .order_by(IngestedDocument.created_at.desc())
        )
        rows = result.all()

        return [
            {
                "id": str(d.id),
                "original_filename": d.original_filename,
                "file_type": d.file_type,
                "document_category": d.document_category,
                "arguider_status": d.arguider_status,
                # DT-022: expõe mensagem de erro na listagem para a UI mostrar
                # o motivo direto no row (sem precisar clicar pra ver detalhe).
                # Só é preenchida quando arguider_status='error'.
                "arguider_error_message": d.arguider_error_message,
                # DT-029: expõe motivo da quarentena na listagem (tipos PII
                # detectados) para o GP poder decidir se é falso-positivo
                # e liberar via botão.
                "quarantine_status": getattr(d, "quarantine_status", "none"),
                "pii_fields": (json.loads(d.pii_fields) if d.pii_fields else []) if getattr(d, "pii_detected", False) else [],
                "ocg_updated": d.ocg_updated,
                "file_size_bytes": d.file_size_bytes,
                "created_at": d.created_at.isoformat() if d.created_at else None,
                "source_type": getattr(d, "source_type", None),
                "source_url": getattr(d, "source_url", None),
                "source_repo_id": str(d.source_repo_id) if getattr(d, "source_repo_id", None) else None,
                "content_status": getattr(d, "content_status", "available"),
                # MVP 8 Fase 1 — feedback de progresso (expostos na listagem
                # pra barra renderizar sem depender de chamada extra ao /status)
                "arguider_stage": getattr(d, "arguider_stage", None),
                "arguider_progress_percent": getattr(d, "arguider_progress_percent", 0),
                "arguider_stage_updated_at": (
                    d.arguider_stage_updated_at.isoformat()
                    if getattr(d, "arguider_stage_updated_at", None) else None
                ),
                # MVP X Fase X.Y — tokens usado na análise Arguidor (custo LLM)
                "tokens_used": tokens_used,
            }
            for d, tokens_used in rows
        ]

    async def get_document_detail(self, project_id: UUID, document_id: UUID) -> dict | None:
        """Documento completo + análise do Arguidor.

        MVP 34: docs soft-deleted retornam None (não acessíveis via UI).
        Endpoint pode chamar `get_document_detail_including_deleted` se
        precisar inspecionar deletados (audit/admin).
        """
        result = await self.db.execute(
            select(IngestedDocument).where(
                IngestedDocument.id == document_id,
                IngestedDocument.project_id == project_id,
                IngestedDocument.deleted_at.is_(None),
            )
        )
        doc = result.scalar_one_or_none()
        if not doc:
            return None

        detail = {
            "id": str(doc.id),
            "original_filename": doc.original_filename,
            "file_type": doc.file_type,
            "document_category": doc.document_category,
            "arguider_status": doc.arguider_status,
            "arguider_error_message": doc.arguider_error_message,
            "ocg_updated": doc.ocg_updated,
            "file_size_bytes": doc.file_size_bytes,
            "created_at": doc.created_at.isoformat() if doc.created_at else None,
        }

        # Buscar análise
        analysis_result = await self.db.execute(
            select(ArguiderAnalysis).where(
                (ArguiderAnalysis.document_id == document_id) &
                (ArguiderAnalysis.project_id == project_id)
            )
        )
        analysis = analysis_result.scalar_one_or_none()
        if analysis:
            detail["analysis"] = {
                "classification": json.loads(analysis.document_classification) if analysis.document_classification else {},
                "gaps": json.loads(analysis.gaps) if analysis.gaps else [],
                "show_stoppers": json.loads(analysis.show_stoppers) if analysis.show_stoppers else [],
                "poor_definitions": json.loads(analysis.poor_definitions) if analysis.poor_definitions else [],
                "improvement_suggestions": json.loads(analysis.improvement_suggestions) if analysis.improvement_suggestions else [],
                "module_candidates": json.loads(analysis.module_candidates) if analysis.module_candidates else [],
                "ocg_fields_to_update": json.loads(analysis.ocg_fields_to_update) if analysis.ocg_fields_to_update else [],
                "tokens_used": analysis.tokens_used,
                "latency_ms": analysis.latency_ms,
            }

        return detail

    async def get_document_status(self, project_id: UUID, document_id: UUID) -> dict | None:
        """Status para polling.

        MVP 34: docs soft-deleted retornam None (frontend para de poll quando recebe 404).
        """
        result = await self.db.execute(
            select(IngestedDocument).where(
                IngestedDocument.id == document_id,
                IngestedDocument.project_id == project_id,
                IngestedDocument.deleted_at.is_(None),
            )
        )
        doc = result.scalar_one_or_none()
        if not doc:
            return None
        return {
            "document_id": str(doc.id),
            "arguider_status": doc.arguider_status,
            "arguider_started_at": doc.arguider_started_at.isoformat() if doc.arguider_started_at else None,
            "arguider_completed_at": doc.arguider_completed_at.isoformat() if doc.arguider_completed_at else None,
            "ocg_updated": doc.ocg_updated,
            # MVP 8 Fase 1 — feedback de progresso
            "arguider_stage": doc.arguider_stage,
            "arguider_progress_percent": doc.arguider_progress_percent,
            "arguider_stage_updated_at": doc.arguider_stage_updated_at.isoformat() if doc.arguider_stage_updated_at else None,
        }

    async def delete_document(
        self,
        project_id: UUID,
        document_id: UUID,
        actor_id: Optional[UUID] = None,
        reason: str = "manual",
    ) -> dict:
        """Soft-delete + reversão de propagação assíncrona (MVP 34).

        ## Breaking change vs. comportamento pré-MVP 34

        Antes (até 2026-05-03): hard-delete síncrono + reversão via
        `_contract_ocg_for_deleted_document` baseado em `ocg_delta_log`.
        Retornava `{"success": True, "ocg_contraction": ...}` com 200 OK.

        Depois (MVP 34): soft-delete imediato + Celery job assíncrono de
        recompute. Retorna `{"success": True, "revert_job_id": ..., "status_code": 202}`.
        Frontend deve fazer polling do status via GET endpoint dedicado.

        ## Validações canônicas (preservadas)

        - Doc deve existir e pertencer ao projeto
        - Doc não pode estar em processamento
        - Doc não pode ter módulos APROVADOS dependentes (bloqueia)

        ## Args

            project_id: UUID do projeto.
            document_id: UUID do doc.
            actor_id: UUID do user que disparou (None = sistema).
            reason: 'manual'|'lgpd'|'smoke_cleanup' (CHECK constraint no DB).

        ## Retorno

            dict com campos canônicos:
              - success (bool)
              - status_code (int) — 202 sucesso, 4xx erro
              - revert_job_id (str) quando 202
              - error (str) quando falha
        """
        from datetime import datetime, timezone

        result = await self.db.execute(
            select(IngestedDocument).where(
                IngestedDocument.id == document_id,
                IngestedDocument.project_id == project_id,
            )
        )
        doc = result.scalar_one_or_none()
        if not doc:
            return {"success": False, "error": "Documento não encontrado", "status_code": 404}

        if doc.deleted_at is not None:
            return {
                "success": False,
                "error": "Documento já foi deletado",
                "status_code": 409,
                "deleted_at": doc.deleted_at.isoformat(),
            }

        if doc.arguider_status == "processing":
            return {"success": False, "error": "Documento em análise, aguarde conclusão", "status_code": 409}

        # Validação canônica preservada: módulos aprovados bloqueiam delete.
        # Cleanup do candidato é responsabilidade do GP (rejeitar antes).
        mc_result = await self.db.execute(
            select(func.count(ModuleCandidate.id)).where(
                ModuleCandidate.project_id == project_id,
                ModuleCandidate.status == "approved",
                ModuleCandidate.source_document_ids.contains(str(document_id)),
            )
        )
        approved_count = mc_result.scalar() or 0
        if approved_count > 0:
            return {"success": False, "error": "Módulos aprovados dependem deste documento", "status_code": 409}

        # Validação do reason (também tem CHECK no DB, mas valida cedo
        # para devolver erro amigável em vez de IntegrityError).
        if reason not in ("manual", "lgpd", "smoke_cleanup"):
            return {
                "success": False,
                "error": f"reason inválido: {reason!r} — esperado manual|lgpd|smoke_cleanup",
                "status_code": 400,
            }

        # Marca soft-delete imediato (antes de enfileirar — Celery worker
        # vê deleted_at IS NOT NULL e idempotência layer 2 detecta).
        doc.deleted_at = datetime.now(timezone.utc)
        doc.deleted_by = actor_id
        doc.deleted_reason = reason
        await self.db.commit()

        # Enfileira Celery task. `revert_document_propagation_task` faz
        # o recompute do OCG + cleanup auxiliar + audit + payload.
        from app.tasks.pipeline import revert_document_propagation_task

        async_result = revert_document_propagation_task.delay(
            document_id=str(document_id),
            project_id=str(project_id),
            actor_id=str(actor_id) if actor_id else None,
            reason=reason,
        )

        logger.info(
            "ingestion.document_soft_deleted_revert_enqueued",
            document_id=str(document_id),
            project_id=str(project_id),
            actor_id=str(actor_id) if actor_id else None,
            reason=reason,
            revert_job_id=async_result.id,
        )

        return {
            "success": True,
            "status_code": 202,
            "revert_job_id": async_result.id,
            "message": "Documento marcado para reversão. Job em background recalcula OCG.",
        }

    async def _contract_ocg_for_deleted_document(
        self, project_id: UUID, document_id: UUID
    ) -> dict:
        """Reverte no OCG os campos alterados pelo doc, exceto os que foram
        tocados por deltas posteriores (evita perder contribuições de outros
        documentos).

        Retorna dict com `fields_reverted` e `fields_skipped`. Se o doc não
        teve deltas, retorna `{"fields_reverted": [], "fields_skipped": []}`.
        """
        # 1. Buscar deltas deste doc em ordem cronológica (mais antigo primeiro).
        doc_deltas_q = await self.db.execute(
            select(OCGDeltaLog).where(
                OCGDeltaLog.project_id == project_id,
                OCGDeltaLog.document_id == document_id,
            ).order_by(OCGDeltaLog.created_at.asc())
        )
        doc_deltas = list(doc_deltas_q.scalars().all())
        if not doc_deltas:
            return {"fields_reverted": [], "fields_skipped": []}

        # 2. Para cada campo tocado por este doc, descobrir o valor "antes"
        #    do primeiro delta que o tocou. O delta mais antigo guarda o `old`
        #    em `fields_changed[field].old` — essa é a origem da reversão.
        first_old: dict[str, any] = {}
        for delta in doc_deltas:
            try:
                changes = json.loads(delta.fields_changed) if delta.fields_changed else {}
            except (ValueError, TypeError):
                changes = {}
            for field, change in changes.items():
                if field not in first_old and isinstance(change, dict) and "old" in change:
                    first_old[field] = change["old"]

        if not first_old:
            return {"fields_reverted": [], "fields_skipped": []}

        touched_fields = set(first_old.keys())
        doc_delta_ids = {d.id for d in doc_deltas}

        # 3. Verificar deltas posteriores (de outros docs ou de propagação) que
        #    tocaram os mesmos campos. Para cada campo com delta posterior,
        #    NÃO reverte — apenas registra no fields_skipped.
        later_q = await self.db.execute(
            select(OCGDeltaLog).where(
                OCGDeltaLog.project_id == project_id,
                OCGDeltaLog.created_at > doc_deltas[0].created_at,
            ).order_by(OCGDeltaLog.created_at.asc())
        )
        later_deltas = [d for d in later_q.scalars().all() if d.id not in doc_delta_ids]

        fields_blocked = set()
        for later in later_deltas:
            try:
                later_changes = json.loads(later.fields_changed) if later.fields_changed else {}
            except (ValueError, TypeError):
                later_changes = {}
            for field in later_changes:
                if field in touched_fields:
                    fields_blocked.add(field)

        fields_to_revert = touched_fields - fields_blocked

        # 4. Carregar OCG atual; se nenhum campo a reverter, só registra delta
        #    documental sem alterar OCG.
        ocg_q = await self.db.execute(
            select(OCG).where(OCG.project_id == project_id).order_by(OCG.created_at.desc()).limit(1)
        )
        ocg = ocg_q.scalar_one_or_none()
        if not ocg:
            return {
                "fields_reverted": [],
                "fields_skipped": sorted(fields_blocked),
                "note": "OCG inexistente; nada a contrair",
            }

        try:
            ocg_current = json.loads(ocg.ocg_data) if ocg.ocg_data else {}
        except (ValueError, TypeError):
            ocg_current = {}

        reverted_changes = {}
        for field in fields_to_revert:
            old_value = first_old[field]
            current_value = ocg_current.get(field)
            ocg_current[field] = old_value
            reverted_changes[field] = {
                "old": current_value,
                "new": old_value,
                "reasoning": f"Revertido por remoção do documento {document_id}",
            }

        # 5. Se houve reversão, incrementar versão + gravar OCG novo + delta.
        version_from = ocg.version
        version_to = version_from + (1 if reverted_changes else 0)

        if reverted_changes:
            ocg.ocg_data = json.dumps(ocg_current, ensure_ascii=False)
            ocg.version = version_to
            ocg.updated_at = datetime.now(timezone.utc)

        # Sempre grava delta de removal (mesmo que fields_to_revert vazio) —
        # audit trail de que o doc foi removido e quais campos ficaram presos.
        summary_parts = []
        if reverted_changes:
            summary_parts.append(
                f"Contraídos {len(reverted_changes)} campo(s): {sorted(reverted_changes)}"
            )
        if fields_blocked:
            summary_parts.append(
                f"Não contraídos (tocados por deltas posteriores): {sorted(fields_blocked)}"
            )
        change_summary = " | ".join(summary_parts) if summary_parts else "Remoção sem impacto no OCG"

        delta = OCGDeltaLog(
            project_id=project_id,
            document_id=None,  # doc está sendo deletado; FK seria inválida
            ocg_version_from=version_from,
            ocg_version_to=version_to,
            fields_changed=json.dumps(reverted_changes, ensure_ascii=False),
            change_summary=change_summary,
            trigger_source="document_removal",
            ocg_snapshot=json.dumps(ocg_current, ensure_ascii=False) if reverted_changes else None,
        )
        self.db.add(delta)

        return {
            "fields_reverted": sorted(reverted_changes.keys()),
            "fields_skipped": sorted(fields_blocked),
            "version_from": version_from,
            "version_to": version_to,
        }
