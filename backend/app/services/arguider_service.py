"""
Arguider Service — Agent 9: Análise de documentos ingeridos.
Classifica, identifica gaps, show-stoppers, má-definição,
sugere módulos e atualiza OCG evolutivamente.
"""
import json
import hashlib
import re
from datetime import datetime, timezone
from typing import Any, Dict, List, Optional

from app.utils.retry import gca_retry
from uuid import uuid4, UUID

from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession
from anthropic import AsyncAnthropic
import structlog

from app.core.config import settings
from app.models.base import (
    IngestedDocument, ArguiderAnalysis, ModuleCandidate,
    OCG,
)

logger = structlog.get_logger(__name__)


# ============================================================================
# JSON parsing helpers (DT-067)
# ============================================================================
_CODE_FENCE_RE = re.compile(
    r"^```(?:json|JSON)?\s*\n?(?P<body>.*?)\n?```\s*$",
    re.DOTALL,
)


def _strip_code_fence(text: str) -> str:
    """Se o texto todo estiver dentro de ```json ... ```, retorna o
    conteúdo interno. Caso contrário, retorna o texto como veio."""
    match = _CODE_FENCE_RE.match(text.strip())
    if match:
        return match.group("body").strip()
    return text


def _extract_balanced_object(text: str) -> Optional[str]:
    """Encontra o primeiro objeto JSON bem-formado com `{}` balanceado
    no texto. Respeita strings (inclusive escapes e aspas internas) pra
    não contar chaves que fazem parte de valores textuais.

    Retorna o trecho exato (substring) ou None quando não há `{` válido.
    """
    start = text.find("{")
    if start < 0:
        return None

    depth = 0
    in_string = False
    escape = False
    for i in range(start, len(text)):
        char = text[i]
        if escape:
            escape = False
            continue
        if char == "\\":
            escape = True
            continue
        if char == '"':
            in_string = not in_string
            continue
        if in_string:
            continue
        if char == "{":
            depth += 1
        elif char == "}":
            depth -= 1
            if depth == 0:
                return text[start : i + 1]
    return None


# ============================================================================
# Document Extractor
# ============================================================================

class DocumentExtractor:  # noqa: E302
    """Extrator multi-formato.

    Quando instanciado com `project_id` + `db`, pode acionar camada 3
    do pipeline de PDF (OCR via LLM Vision do provider configurado no
    projeto). Sem esses args, apenas camadas 1+2 do PDF ficam ativas —
    útil pra testes e extração offline.
    """
    """Extrai texto de diferentes tipos de arquivo."""

    def __init__(self, client=None, project_id=None, db=None):
        self.client = client
        self.project_id = project_id
        self.db = db

    async def extract_text(self, file_bytes: bytes, file_type: str) -> str:
        if file_type in ("markdown", "md", "txt", "code"):
            return file_bytes.decode("utf-8", errors="replace")

        if file_type == "pdf":
            # MVP 8 Fase 3B — se temos contexto de projeto, ativa OCR via
            # Vision como camada 3 (quando 1+2 não produzem texto).
            if self.project_id is not None and self.db is not None:
                return await self._extract_pdf_with_vision_ocr(file_bytes)
            return self._extract_pdf(file_bytes)

        if file_type in ("docx", "doc"):
            return self._extract_docx(file_bytes)

        if file_type in ("xlsx", "xls", "csv", "spreadsheet"):
            return self._extract_spreadsheet(file_bytes, file_type)

        if file_type in ("image", "wireframe", "png", "jpg", "jpeg", "gif", "webp"):
            return await self._extract_image_description(file_bytes)

        # Tentar como texto
        try:
            return file_bytes.decode("utf-8", errors="replace")
        except Exception:
            return "{arquivo binário não extraível}"

    def _extract_pdf(self, file_bytes: bytes) -> str:
        """MVP 8 Fase 3 Commit A — delega ao pipeline de camadas que
        tenta AcroForm + texto pesquisável (e, no Commit B, OCR via
        Vision). Sem isso, formulários PDF ficavam invisíveis e
        escaneados retornavam string vazia silenciosamente."""
        try:
            from app.services.pdf_layered_extractor import extract_pdf_layered
            result = extract_pdf_layered(file_bytes)
            if result.text:
                return result.text
            # Sem camada produtiva: preserva aviso pro log do pipeline
            warnings = " / ".join(result.warnings) if result.warnings else "sem conteúdo extraído"
            logger.warning("extractor.pdf_empty_layers", warnings=warnings)
            return f"[PDF sem texto extraível — {warnings}]"
        except Exception as e:
            logger.warning("extractor.pdf_error", error=str(e))
            return f"[Erro ao extrair PDF: {str(e)}]"

    async def _extract_pdf_with_vision_ocr(self, file_bytes: bytes) -> str:
        """MVP 8 Fase 3B — pipeline completo com camadas 1+2+3.

        Camadas 1 e 2 (AcroForm + texto pesquisável) rodam síncronas
        como no path sem contexto. Só se nenhuma delas produzir texto,
        aciona OCR via `vision_service` usando o provider do projeto.

        OCR custa tokens — nunca é disparado por precaução, só quando
        o PDF é realmente escaneado/imagem e o projeto tem provider
        com visão configurado.
        """
        try:
            from app.services.pdf_layered_extractor import extract_pdf_layered_with_ocr
            from app.services.vision_service import ocr_pdf_via_project_vision

            async def _ocr_cb(pdf_bytes: bytes):
                return await ocr_pdf_via_project_vision(
                    pdf_bytes, self.db, self.project_id,
                )

            result = await extract_pdf_layered_with_ocr(file_bytes, ocr_callback=_ocr_cb)
            if result.text:
                return result.text
            warnings = " / ".join(result.warnings) if result.warnings else "sem conteúdo extraído"
            logger.warning(
                "extractor.pdf_empty_layers_after_ocr",
                project_id=str(self.project_id),
                warnings=warnings,
            )
            return f"[PDF sem texto extraível após OCR — {warnings}]"
        except Exception as e:
            logger.warning(
                "extractor.pdf_vision_error",
                project_id=str(self.project_id) if self.project_id else None,
                error=str(e),
            )
            return f"[Erro ao extrair PDF com OCR: {str(e)}]"

    def _extract_docx(self, file_bytes: bytes) -> str:
        """MVP 8 Fase 2 — delega ao rich extractor que percorre tabelas,
        caixas de texto, headers/footers e notas de rodapé, não só
        doc.paragraphs. Sem isso, documentos com RFs em tabela viram
        texto vazio pro Arguidor e o OCG não evolui."""
        try:
            from app.services.rich_docx_extractor import extract_rich_text
            return extract_rich_text(file_bytes)
        except Exception as e:
            logger.warning("extractor.docx_error", error=str(e))
            return f"[Erro ao extrair DOCX: {str(e)}]"

    def _extract_spreadsheet(self, file_bytes: bytes, file_type: str) -> str:
        if file_type == "csv":
            return file_bytes.decode("utf-8", errors="replace")
        try:
            import io
            from openpyxl import load_workbook
            wb = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
            rows = []
            for ws in wb.worksheets[:5]:
                rows.append(f"=== Aba: {ws.title} ===")
                for row in ws.iter_rows(max_row=200, values_only=True):
                    cells = [str(c) if c is not None else "" for c in row]
                    if any(cells):
                        rows.append(" | ".join(cells))
            return "\n".join(rows)
        except ImportError:
            return "[openpyxl não instalado]"
        except Exception as e:
            return f"[Erro ao extrair planilha: {str(e)}]"

    async def _extract_image_description(self, image_bytes: bytes) -> str:
        """Usa Claude Vision para descrever imagem/wireframe."""
        try:
            import base64
            client = self.client  # Reutiliza client com chave do projeto
            b64 = base64.b64encode(image_bytes).decode("ascii")
            response = await client.messages.create(
                model=settings.ANTHROPIC_MODEL,
                max_tokens=2048,
                messages=[{
                    "role": "user",
                    "content": [
                        {
                            "type": "image",
                            "source": {"type": "base64", "media_type": "image/png", "data": b64},
                        },
                        {
                            "type": "text",
                            "text": "Descreva detalhadamente este wireframe/imagem de interface ou diagrama. "
                                    "Identifique: elementos visuais, fluxos, componentes, textos visíveis e funcionalidades implícitas.",
                        },
                    ],
                }],
            )
            return response.content[0].text
        except Exception as e:
            logger.warning("extractor.vision_error", error=str(e))
            return f"[Erro ao analisar imagem via IA: {str(e)}]"


# ============================================================================
# Arguider Service (Agent 9)
# ============================================================================

ARGUIDER_SYSTEM_PROMPT = """
Você é o Arguidor do GCA (Gestão de Codificação Assistida).
Seu papel é analisar documentos ingeridos em projetos de software e:

1. Classificar o tipo e categoria do documento
2. Identificar GAPS em relação ao OCG (Objeto Contexto Global) do projeto
3. Identificar SHOW-STOPPERS: contradições graves que impedem implementação
4. Identificar MÁ DEFINIÇÃO: ambiguidades que precisam ser esclarecidas
5. Sugerir MELHORIAS de forma objetiva e acionável
6. Identificar MÓDULOS CANDIDATOS: funcionalidades implementáveis

Para cada módulo candidato, decida se é:
- 'feature': funcionalidade completa de negócio
- 'component': componente técnico reutilizável

IMPORTANTE:
- Seja específico. Cada item com ID único (G001, SS001, PD001, IS001).
- Módulo só é ready_for_codegen=true se o documento fornece TODAS as informações.
- Ao atualizar OCG, sugira apenas campos diretamente impactados.
- Responda SOMENTE com JSON válido.
"""


class ArguiderService:
    """Agent 9: Arguidor — analisa documentos e atualiza OCG.

    CAMADA PROJETO — usa chave de IA configurada pelo GP do projeto.
    Criticidade (contrato §6.2): **ALTA** (análise que alimenta o OCG e dispara
    módulos candidatos e achados). Sem fallback silencioso para chave global
    do admin — se o projeto não tem chave, levanta `RuntimeError` claro.
    """

    def __init__(
        self,
        db: AsyncSession,
        project_api_key: str = None,
        provider: str = None,
        model: str = None,
        base_url: str = None,
    ):
        """Arguidor multi-provider (DT-032 + DT-023).

        Aceita:
          - `project_api_key`: chave do projeto. Obrigatória para todos
            os providers exceto **ollama** (local, sem auth por default).
          - `provider`: "anthropic"|"openai"|"deepseek"|"grok"|"gemini"|"ollama".
            Default `"anthropic"` para retrocompat com callers antigos.
          - `model`: modelo específico. Se None, usa default do provider.
          - `base_url`: **obrigatório quando provider=ollama**. Endpoint
            do daemon Ollama do GP (ex: http://host.docker.internal:11434).
            Ignorado pelos demais providers.

        Contrato §6.4: nenhum fallback para chave global do admin.
        """
        self.db = db
        self.provider = (provider or "anthropic").lower()
        is_ollama = self.provider == "ollama"

        if not is_ollama and not project_api_key:
            raise RuntimeError(
                "Arguidor não pode operar sem chave IA do projeto. "
                "GP deve configurar provedor e chave em Settings > LLM. "
                "Arguidor é tarefa de ALTA criticidade (contrato §6.2) e "
                "não aceita fallback para chave global do admin."
            )
        if is_ollama and not base_url:
            raise RuntimeError(
                "Arguidor com provider=ollama exige `base_url`. "
                "GP deve informar o endpoint do daemon Ollama em Settings > LLM "
                "(ex: http://host.docker.internal:11434)."
            )
        self.api_key = project_api_key  # pode ser None pro ollama
        self.base_url = (base_url or "").rstrip("/") if base_url else None
        # Model default por provider (match com o que `/settings/llm/validate`
        # testa em settings_router.py).
        _default_models = {
            "anthropic": "claude-haiku-4-5-20251001",
            "openai": "gpt-4o-mini",
            "deepseek": "deepseek-chat",
            "grok": "grok-2",
            "gemini": "gemini-2.0-flash",
            "ollama": "llama3.1:8b",
        }
        self.model = model or _default_models.get(self.provider, "deepseek-chat")
        # Client Anthropic só quando for o provider — para o resto, httpx.
        self.client = AsyncAnthropic(api_key=project_api_key) if self.provider == "anthropic" else None
        # DocumentExtractor ainda recebe o client Anthropic por compatibilidade
        # com o pipeline de extração de texto. Quando provider é outro, passa
        # None — extração fallback usa pypdf local sem LLM.
        self.extractor = DocumentExtractor(client=self.client)

    async def _call_llm(
        self,
        system_prompt: str,
        user_prompt: str,
        max_tokens: int = 4096,
    ) -> tuple[str, int]:
        """Chamada unificada ao LLM do projeto (DT-032).

        Replica o padrão de `agent_service._call_llm` + `ocg_updater_service`:
          - Anthropic: SDK nativo AsyncAnthropic.
          - OpenAI/DeepSeek/Grok: httpx POST em `/v1/chat/completions`
            (endpoint OpenAI-compatible).
          - Gemini: escopo futuro (requer formato próprio, fora da DT-032
            que mira o caminho já usado pelo user dogfood).

        Retorna `(text, tokens_used)`.
        """
        if self.provider == "anthropic" and self.client:
            response = await self.client.messages.create(
                model=self.model,
                max_tokens=max_tokens,
                temperature=0.2,
                system=system_prompt,
                messages=[{"role": "user", "content": user_prompt}],
            )
            text = response.content[0].text
            tokens = response.usage.input_tokens + response.usage.output_tokens
            # DT-069 — detectar truncamento por limite de tokens. Sem isso,
            # o JSON vem cortado no meio, o parse falha e o Arguidor
            # persiste arrays vazios silenciosamente. Levanta pra caller
            # decidir (em vez de logar warning e seguir com texto truncado).
            stop_reason = getattr(response, "stop_reason", None)
            if stop_reason == "max_tokens":
                raise RuntimeError(
                    f"LLM truncou resposta por max_tokens={max_tokens}. "
                    f"Aumente ANTHROPIC_MAX_TOKENS no .env ou reduza o "
                    f"tamanho do documento. Modelo: {self.model}, "
                    f"tokens consumidos: {tokens}."
                )
            return text, tokens

        import httpx
        # DT-023: Ollama usa endpoint OpenAI-compatible (`/v1/chat/completions`)
        # do daemon local. URL base vem do project_settings (campo base_url).
        provider_urls = {
            "openai": "https://api.openai.com/v1/chat/completions",
            "deepseek": "https://api.deepseek.com/chat/completions",
            "grok": "https://api.x.ai/v1/chat/completions",
            "ollama": f"{self.base_url}/v1/chat/completions" if self.base_url else None,
        }
        url = provider_urls.get(self.provider)
        if not url:
            raise ValueError(
                f"Provider '{self.provider}' ainda não suportado pelo Arguidor. "
                f"Suportados hoje: anthropic, openai, deepseek, grok, ollama."
            )

        # Ollama típico não exige Authorization. Caller só inclui Bearer
        # se houver api_key (ex: reverse proxy na frente do daemon).
        headers = {"Content-Type": "application/json"}
        if self.api_key:
            headers["Authorization"] = f"Bearer {self.api_key}"
        async with httpx.AsyncClient(timeout=120.0) as client:
            resp = await client.post(
                url,
                headers=headers,
                json={
                    "model": self.model,
                    "max_tokens": max_tokens,
                    "temperature": 0.2,
                    "messages": [
                        {"role": "system", "content": system_prompt},
                        {"role": "user", "content": user_prompt},
                    ],
                },
            )
        if resp.status_code not in (200, 201):
            raise ValueError(
                f"LLM API error ({resp.status_code}) no provider {self.provider}: "
                f"{resp.text[:300]}"
            )
        data = resp.json()
        text = data["choices"][0]["message"]["content"]
        usage = data.get("usage", {})
        tokens = usage.get("prompt_tokens", 0) + usage.get("completion_tokens", 0)
        return text, tokens

    @gca_retry()
    async def analyze_document(
        self,
        document_id: UUID,
        project_id: UUID,
        document_text: str,
        current_ocg: dict,
        previous_analyses: list[dict] | None = None,
    ) -> dict:
        """Executa análise completa do Arguidor para um documento."""
        try:
            # Atualizar status
            doc = await self.db.execute(
                select(IngestedDocument).where(IngestedDocument.id == document_id)
            )
            document = doc.scalar_one_or_none()
            if document:
                document.arguider_status = "processing"
                document.arguider_started_at = datetime.now(timezone.utc)
                await self.db.commit()

            # Construir prompt
            user_prompt = self._build_prompt(document_text, current_ocg, previous_analyses or [])

            # Chamar LLM do projeto (multi-provider, DT-032)
            start_time = datetime.now(timezone.utc)
            response_text, tokens_used = await self._call_llm(
                system_prompt=ARGUIDER_SYSTEM_PROMPT,
                user_prompt=user_prompt,
                max_tokens=settings.ANTHROPIC_MAX_TOKENS,
            )
            latency_ms = int((datetime.now(timezone.utc) - start_time).total_seconds() * 1000)

            # Parsear JSON
            result_json = self._extract_json(response_text)

            # Salvar análise
            analysis = ArguiderAnalysis(
                document_id=document_id,
                project_id=project_id,
                document_classification=json.dumps(result_json.get("document_classification", {}), ensure_ascii=False),
                gaps=json.dumps(result_json.get("gaps", []), ensure_ascii=False),
                show_stoppers=json.dumps(result_json.get("show_stoppers", []), ensure_ascii=False),
                poor_definitions=json.dumps(result_json.get("poor_definitions", []), ensure_ascii=False),
                improvement_suggestions=json.dumps(result_json.get("improvement_suggestions", []), ensure_ascii=False),
                module_candidates=json.dumps(result_json.get("module_candidates", []), ensure_ascii=False),
                ocg_fields_to_update=json.dumps(result_json.get("ocg_fields_to_update", []), ensure_ascii=False),
                # DT-032: label reflete o modelo real usado, não mais hardcoded
                # ANTHROPIC_MODEL. Ex: "deepseek-chat", "claude-haiku-4-5-*".
                llm_model=f"{self.provider}:{self.model}",
                tokens_used=tokens_used,
                latency_ms=latency_ms,
            )
            self.db.add(analysis)
            await self.db.commit()

            # Promover module_candidates
            # MVP 9 Fase 9.1 — normaliza module_type pra uma das 6
            # categorias canônicas. LLM pode emitir variações ou valores
            # legados (`component`); o normalizador força retrocompat.
            # MVP 9 Fase 9.1.2 — status inicial é `sugerido` (pt-BR canônico)
            # pro item de novo candidato; ciclo de vida só avança quando
            # GP interage (aguardando_resposta → adicionado → concluido).
            from app.constants.module_categories import (
                normalize_module_type, DEFAULT_MODULE_STATUS,
            )
            for mc in result_json.get("module_candidates", []):
                raw_type = mc.get("module_type", "feature")
                canonical_type = normalize_module_type(raw_type)
                candidate = ModuleCandidate(
                    project_id=project_id,
                    arguider_analysis_id=analysis.id,
                    name=mc.get("name", "Unnamed"),
                    description=mc.get("description", ""),
                    module_type=canonical_type,
                    priority=mc.get("priority", "medium"),
                    status=DEFAULT_MODULE_STATUS,  # pt-BR: "sugerido"
                    dependencies=json.dumps(mc.get("dependencies", []), ensure_ascii=False),
                    source_document_ids=json.dumps([str(document_id)], ensure_ascii=False),
                    pillar_impact=json.dumps(
                        {f"p{i}": f"P{i}" in mc.get("pillar_impact", []) for i in range(1, 8)},
                        ensure_ascii=False,
                    ),
                    ready_for_codegen=mc.get("ready_for_codegen", False),
                )
                self.db.add(candidate)

            # DT-070 — propagar gaps/show_stoppers/poor_definitions/
            # improvement_suggestions pra tabela `gatekeeper_items`, que é
            # a fonte lida pela UI do Arguidor (via /projects/:id/gatekeeper).
            # Sem isso, análise existe em arguider_analyses mas a UI
            # permanece vazia ("Nenhum item pendente do Gatekeeper").
            from app.models.base import GatekeeperItem
            gk_buckets = (
                ("gap", result_json.get("gaps", []) or []),
                ("show_stopper", result_json.get("show_stoppers", []) or []),
                ("poor_definition", result_json.get("poor_definitions", []) or []),
                ("improvement", result_json.get("improvement_suggestions", []) or []),
            )
            prefix_map = {
                "gap": "G", "show_stopper": "SS",
                "poor_definition": "PD", "improvement": "IS",
            }
            for item_type, items in gk_buckets:
                prefix = prefix_map[item_type]
                for idx, item in enumerate(items, start=1):
                    raw_id = item.get("id") if isinstance(item, dict) else None
                    item_id = raw_id or f"{prefix}{idx:03d}"
                    self.db.add(GatekeeperItem(
                        project_id=project_id,
                        arguider_analysis_id=analysis.id,
                        item_type=item_type,
                        item_id_in_analysis=str(item_id)[:10],
                        item_data=json.dumps(item, ensure_ascii=False),
                        status="pending",
                    ))

            # OCG é atualizado em seguida pelo OCGUpdaterService (chamado por
            # ingestion_service). Aqui só sinalizamos a intenção do Arguidor de
            # atualizar — o flag final ocg_updated reflete proposta, não persistência.
            ocg_fields = result_json.get("ocg_fields_to_update", [])
            ocg_updated = bool(ocg_fields)

            # Atualizar categoria do documento
            classification = result_json.get("document_classification", {})
            if document:
                document.document_category = classification.get("category", "other")
                document.arguider_status = "completed"
                document.arguider_completed_at = datetime.now(timezone.utc)
                document.ocg_updated = ocg_updated

            await self.db.commit()

            logger.info(
                "arguider.analysis_complete",
                document_id=str(document_id),
                category=classification.get("category"),
                gaps=len(result_json.get("gaps", [])),
                show_stoppers=len(result_json.get("show_stoppers", [])),
                modules=len(result_json.get("module_candidates", [])),
                ocg_updated=ocg_updated,
                tokens_used=tokens_used,
                latency_ms=latency_ms,
            )

            return result_json

        except Exception as e:
            logger.error("arguider.analysis_error", document_id=str(document_id), error=str(e))
            # Marcar como erro
            doc = await self.db.execute(
                select(IngestedDocument).where(IngestedDocument.id == document_id)
            )
            document = doc.scalar_one_or_none()
            if document:
                document.arguider_status = "error"
                document.arguider_error_message = str(e)[:500]
                await self.db.commit()
            raise

    def _build_prompt(self, doc_text: str, ocg: dict, prev: list) -> str:
        prev_summary = "Nenhuma análise anterior."
        if prev:
            prev_summary = "\n".join(
                f"- Doc {i+1}: {a.get('document_classification', {}).get('category', '?')} "
                f"({len(a.get('gaps', []))} gaps, {len(a.get('module_candidates', []))} módulos)"
                for i, a in enumerate(prev)
            )

        # Truncar texto muito longo
        max_chars = 15000
        if len(doc_text) > max_chars:
            doc_text = doc_text[:max_chars] + f"\n\n[... documento truncado, {len(doc_text)} chars total]"

        return f"""=== DOCUMENTO A ANALISAR ===
{doc_text}

=== OCG ATUAL DO PROJETO ===
{json.dumps(ocg, ensure_ascii=False, indent=2)[:5000]}

=== ANÁLISES ANTERIORES (RESUMO) ===
{prev_summary}

=== INSTRUÇÕES ===
Analise o documento acima em relação ao OCG do projeto.
Retorne SOMENTE JSON válido (sem ``` code fences, sem preâmbulo) com as chaves:
document_classification, gaps, show_stoppers, poor_definitions,
improvement_suggestions, module_candidates, ocg_fields_to_update.

DT-069 — LIMITE DE VERBOSIDADE (evita truncamento do LLM):
- `gaps`, `show_stoppers`, `poor_definitions`: máximo 10 itens cada; cada
  `text`/`description` até 300 chars; sem campos narrativos extras além
  do schema.
- `improvement_suggestions`: máximo 8 itens.
- `module_candidates`: máximo 20 itens TOTAIS (ver MVP 9 abaixo).
- `document_classification`: máximo 500 chars total.
- Se houver mais achados do que o limite, priorize os mais críticos e
  cite no `summary` que existem outros não listados.

MVP 9 Fase 9.1 — CATEGORIAS CANÔNICAS DE MÓDULOS (obrigatório):
Cada item de `module_candidates` DEVE ter `module_type` em UMA das 6
categorias canônicas:

  1. `infrastructure`     — Docker/K8s/IaC, provisionamento, rede,
                            secrets storage, volumes persistentes.
  2. `observability`      — Métricas, traces, logs estruturados,
                            dashboards, alertas, health checks.
  3. `middleware`         — Auth/RBAC, rate-limit, CORS, request
                            logging, validação, tratamento de erro.
  4. `backend_service`    — API endpoints, jobs async, workers,
                            integração com DB/cache/fila, conectores
                            externos (sem UI).
  5. `feature`            — Funcionalidade de negócio com valor direto
                            pro usuário final (tela + backend + regras).
  6. `deploy_pipeline`    — CI/CD, testes automatizados em pipeline,
                            release, rollback, migrations gates.

Regras duras:
- Distribua os até 20 itens cobrindo as 6 categorias quando o OCG
  indicar cada camada (ex: se `STACK_RECOMMENDATION.backend.enabled=true`,
  inclua `backend_service` e `middleware`; se `ARCHITECTURE_OVERVIEW.
  execution_model` inclui "Containerizado", inclua `infrastructure` e
  `deploy_pipeline`).
- NÃO devolva `module_type='component'` nem outros valores fora da
  lista — o backend rejeita e força `feature` por default.
- Preencha `description` com 1-2 frases técnicas sobre o QUE o módulo
  faz e QUE PARTE do sistema cobre — não repita o `name`."""

    @staticmethod
    def _extract_json(text: str) -> dict:
        """Extrai objeto JSON da resposta do LLM tolerando formatos comuns:

        1. JSON puro: `{"k": "v"}`
        2. Dentro de code fence: ```json\n{...}\n``` (Claude Haiku costuma
           entregar assim mesmo quando o prompt pede "só JSON").
        3. JSON precedido/seguido por preâmbulo em linguagem natural.
        4. Objeto JSON com chaves `{}` aninhadas — usa contagem de chaves
           em vez de regex gananciosa.

        Quando todas as estratégias falham, loga o texto completo (truncado
        a 2000 chars) em vez de só os primeiros 200. Sem isso, quando a
        análise sai vazia (dogfood 2026-04-19: Haiku respondeu com fence
        e regex `\\{.*\\}` deu match mas parse falhou por conteúdo após
        o primeiro `{`), o operador não consegue diagnosticar.
        """
        if not text:
            logger.warning("arguider.json_parse_failed", reason="empty_text")
            return {}

        stripped = text.strip()
        try:
            return json.loads(stripped)
        except json.JSONDecodeError:
            pass

        fenced = _strip_code_fence(stripped)
        if fenced != stripped:
            try:
                return json.loads(fenced)
            except json.JSONDecodeError:
                stripped = fenced

        extracted = _extract_balanced_object(stripped)
        if extracted is not None:
            try:
                return json.loads(extracted)
            except json.JSONDecodeError as exc:
                logger.warning(
                    "arguider.json_parse_failed",
                    reason="balanced_object_invalid",
                    error=str(exc),
                    extracted_preview=extracted[:500],
                )

        logger.warning(
            "arguider.json_parse_failed",
            reason="no_valid_json_found",
            text_len=len(text),
            text_head=text[:500],
            text_tail=text[-500:],
        )
        return {}
